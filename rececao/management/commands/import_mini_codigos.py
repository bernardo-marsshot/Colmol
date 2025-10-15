from django.core.management.base import BaseCommand
import openpyxl
from rececao.models import MiniCodigo


class Command(BaseCommand):
    help = 'Importa mini códigos do ficheiro Excel'

    def add_arguments(self, parser):
        parser.add_argument('excel_file', type=str, help='Caminho do ficheiro Excel')

    def handle(self, *args, **options):
        excel_file = options['excel_file']
        
        self.stdout.write(f"📂 Abrindo ficheiro: {excel_file}")
        wb = openpyxl.load_workbook(excel_file)
        ws = wb.active
        
        # Linha 1: headers descritivos (ignorar)
        # Linha 2: headers reais (Familia, Mini Codigo, Referencia, Designacao, Identificador, Tipo)
        # Linhas 3+: dados
        
        imported = 0
        updated = 0
        skipped = 0
        
        for row_num in range(3, ws.max_row + 1):
            row = ws[row_num]
            
            familia = row[0].value or ''
            mini_codigo = row[1].value
            referencia = row[2].value or ''
            designacao = row[3].value or ''
            identificador = row[4].value
            tipo = row[5].value or ''
            
            # Validar que mini_codigo existe (campo obrigatório)
            if not mini_codigo or not str(mini_codigo).strip():
                self.stdout.write(self.style.WARNING(
                    f"  ⚠️  Linha {row_num}: Mini código vazio - ignorado"
                ))
                skipped += 1
                continue
            
            # Limpar espaços
            familia = str(familia).strip() if familia else ''
            mini_codigo = str(mini_codigo).strip()
            referencia = str(referencia).strip() if referencia else ''
            designacao = str(designacao).strip() if designacao else ''
            identificador = str(identificador).strip() if identificador else None
            tipo = str(tipo).strip() if tipo else ''
            
            # Criar ou atualizar
            obj, created = MiniCodigo.objects.update_or_create(
                mini_codigo=mini_codigo,
                defaults={
                    'familia': familia,
                    'referencia': referencia,
                    'designacao': designacao,
                    'identificador': identificador,
                    'tipo': tipo,
                }
            )
            
            if created:
                imported += 1
                if imported % 100 == 0:
                    self.stdout.write(f"  ✅ {imported} códigos importados...")
            else:
                updated += 1
        
        self.stdout.write(self.style.SUCCESS(
            f"\n✅ Importação concluída!"
        ))
        self.stdout.write(f"   📊 Novos: {imported}")
        self.stdout.write(f"   🔄 Atualizados: {updated}")
        self.stdout.write(f"   ⚠️  Ignorados: {skipped}")
        self.stdout.write(f"   📈 Total na BD: {MiniCodigo.objects.count()}")
