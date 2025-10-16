# rececao/services.py
import hashlib
import json
import os
import re
import base64
from io import BytesIO
from PIL import Image
import signal
from decimal import Decimal

import PyPDF2
import pytesseract
from pdf2image import convert_from_path
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill

from django.conf import settings
from django.http import HttpResponse
from django.db import transaction

from .models import (InboundDocument, ReceiptLine, CodeMapping, MatchResult,
                     ExceptionTask, POLine, PurchaseOrder)

# PyMuPDF for better PDF text extraction
try:
    import fitz  # PyMuPDF
    PYMUPDF_AVAILABLE = True
    print("✅ PyMuPDF disponível para extração avançada de PDF")
except ImportError:
    PYMUPDF_AVAILABLE = False
    print("⚠️ PyMuPDF não disponível")

# --- QR code detection (usando OpenCV) ---
try:
    import cv2
    import numpy as np
    QR_CODE_ENABLED = True
    print("✅ QR code detection disponível (OpenCV)")
except ImportError:
    QR_CODE_ENABLED = False
    print("⚠️ QR code não disponível (instale opencv-python para ativar)")

# Se precisares especificar o caminho do tesseract no Windows:
# pytesseract.pytesseract.tesseract_cmd = r"C:\Program Files\Tesseract-OCR\tesseract.exe"

# --- Normalização de números (3 casas decimais = milhares) ---
def normalize_number(value_str: str) -> float:
    """
    Normaliza valores numéricos com vírgula, detectando formato de milhares vs decimais.
    
    Regras baseadas no NÚMERO DE DÍGITOS após a vírgula:
    - 3 dígitos após vírgula → formato de milhares (remover vírgula)
      Exemplos: "1,880" → 1880.0, "0,150" → 150.0, "2,000" → 2000.0, "125,000" → 125000.0
    - 1-2 dígitos após vírgula → formato decimal (substituir vírgula por ponto)
      Exemplos: "1,88" → 1.88, "2,5" → 2.5, "34,00" → 34.0
    
    Args:
        value_str: String com número (pode ter vírgula)
    
    Returns:
        float: Valor numérico convertido corretamente
    """
    if not value_str or not isinstance(value_str, str):
        return 0.0
    
    # Remover espaços
    value_str = value_str.strip()
    
    # Se não tem vírgula, converter diretamente
    if ',' not in value_str:
        try:
            return float(value_str.replace(' ', ''))
        except ValueError:
            return 0.0
    
    # Detectar quantas casas decimais após a vírgula
    parts = value_str.split(',')
    if len(parts) != 2:
        # Múltiplas vírgulas ou formato inválido
        try:
            return float(value_str.replace(',', '').replace(' ', ''))
        except ValueError:
            return 0.0
    
    integer_part = parts[0].replace(' ', '')
    decimal_part = parts[1].replace(' ', '')
    
    # Se tem exatamente 3 dígitos após vírgula → formato de milhares
    # Remover vírgula completamente: "1,880" → "1880", "2,000" → "2000"
    if len(decimal_part) == 3:
        try:
            return float(integer_part + decimal_part)
        except ValueError:
            return 0.0
    
    # Caso contrário (1-2 dígitos) → formato decimal normal
    # Substituir vírgula por ponto: "2,5" → "2.5"
    try:
        return float(f"{integer_part}.{decimal_part}")
    except ValueError:
        return 0.0

# --- OCR.space API (Level 0 - Cloud OCR com 25k req/mês grátis) ---
try:
    import requests
    OCR_SPACE_AVAILABLE = True
except ImportError:
    OCR_SPACE_AVAILABLE = False

def ocr_space_api(file_path: str, language='por'):
    """
    OCR.space API - Level 0 (prioridade máxima)
    - 25.000 requisições/mês grátis
    - Suporta PT, ES, FR, EN + 30 idiomas
    - Detecção automática de tabelas
    - Fallback: retorna None se falhar
    """
    if not OCR_SPACE_AVAILABLE:
        return None
    
    api_key = os.environ.get('OCR_SPACE_API_KEY')
    if not api_key:
        print("⚠️ OCR_SPACE_API_KEY não encontrada - usando engines locais")
        return None
    
    try:
        url = 'https://api.ocr.space/parse/image'
        
        # Mapeamento de idiomas (por=português, spa=espanhol, fre=francês)
        lang_map = {'por': 'por', 'pt': 'por', 'es': 'spa', 'spa': 'spa', 'fr': 'fre', 'fre': 'fre', 'en': 'eng'}
        ocr_language = lang_map.get(language.lower(), 'por')
        
        with open(file_path, 'rb') as f:
            payload = {
                'apikey': api_key,
                'language': ocr_language,
                'isOverlayRequired': False,
                'detectOrientation': True,
                'scale': True,
                'OCREngine': 2,  # Engine 2 é mais preciso para tabelas
                'isTable': True  # Detecção de tabelas ativada
            }
            
            response = requests.post(url, files={'file': f}, data=payload, timeout=30)
            
            if response.status_code == 200:
                result = response.json()
                
                if result.get('IsErroredOnProcessing'):
                    print(f"⚠️ OCR.space error: {result.get('ErrorMessage', 'Unknown error')}")
                    return None
                
                # Extrai texto de todas as páginas
                text_parts = []
                if result.get('ParsedResults'):
                    for page in result['ParsedResults']:
                        page_text = page.get('ParsedText', '')
                        if page_text:
                            text_parts.append(page_text)
                
                full_text = '\n'.join(text_parts)
                
                if full_text.strip():
                    print(f"✅ OCR.space (API): {len(full_text)} chars extraídos")
                    return full_text
                else:
                    print("⚠️ OCR.space retornou texto vazio - fallback para engines locais")
                    return None
            else:
                print(f"⚠️ OCR.space HTTP {response.status_code} - fallback para engines locais")
                return None
                
    except requests.Timeout:
        print("⚠️ OCR.space timeout (30s) - fallback para engines locais")
        return None
    except Exception as e:
        print(f"⚠️ OCR.space exception: {e} - fallback para engines locais")
        return None

# --- Imports opcionais para extração universal ---
try:
    import pdfplumber
    PDFPLUMBER_AVAILABLE = True
except ImportError:
    PDFPLUMBER_AVAILABLE = False

try:
    import camelot
    CAMELOT_AVAILABLE = True
except ImportError:
    CAMELOT_AVAILABLE = False

try:
    from rapidfuzz import fuzz, process
    RAPIDFUZZ_AVAILABLE = True
except ImportError:
    RAPIDFUZZ_AVAILABLE = False

# --- LLM para Document Extraction (Groq + Ollama) ---

def groq_extract_document(file_path: str, ocr_text: str, api_key: str, key_name: str = "GROQ_API_KEY"):
    """
    Groq LLM Document Extractor (gratuito, sem instalação)
    Usa Llama-3.3-70B para extrair dados estruturados
    
    Returns:
        tuple: (extracted_data, status_code) ou (None, status_code) se falhar
    """
    try:
        system_prompt = """You are a document extraction expert. Extract ALL product data from invoices, delivery notes, and purchase orders in Portuguese, Spanish, or French.

CRITICAL: Extract EVERY product line, even if incomplete or malformed.

Return valid JSON:
{
  "fornecedor": "supplier name or null",
  "nif": "tax ID or null",
  "numero_documento": "document number or null",
  "data_documento": "YYYY-MM-DD or null",
  "numero_encomenda": "PO number or null",
  "produtos": [
    {
      "codigo": "product code",
      "descricao": "description",
      "quantidade": 10.5,
      "preco_unitario": 25.99,
      "total": 272.40,
      "numero_encomenda": "PO number for this specific product (if document has multiple POs)"
    }
  ]
}

QUANTITY EXTRACTION RULES (CRITICAL):
⚠️ Elastron invoices ONLY (detect by small INTEGER after unit):
  - Pattern: [decimal] [UNIT] [INTEGER ≤3] [price]
  - Example: "34,00 ML 1 3,99" → quantidade=34.0 (NOT 1)
  - The INTEGER after unit (1, 2, or 3) = volume count (number of rolls) → IGNORE
  - The DECIMAL before unit = quantity in meters/units → USE THIS
  - This rule applies ONLY when: number after unit is INTEGER ≤3 (volumes are 1, 2, or 3)

⚠️ Eurospuma/other invoices (decimals after unit = dimensions, not volumes):
  - Pattern: [quantity] [UNIT] [decimal dimensions...]
  - Example: "125,000 UN 1,880 0,150 0,080" → quantidade=125.0 (NOT 1.880)
  - Example: "640,000 MT 320,000 0,260" → quantidade=640.0 (NOT 320.0)
  - Numbers after unit are DECIMALS (dimensions/measurements) → DO NOT confuse with quantity
  - The quantity is the FIRST number before the unit

⚠️ Detection logic:
  - If pattern is [decimal] [unit] [INTEGER 1-3] [price] → Elastron format, ignore integer
  - If pattern is [decimal] [unit] [DECIMAL >3 OR with comma] → Normal format, use first decimal as quantity
  - If just [quantity] [unit] [price] → Standard format

NUMBER FORMATTING (CRITICAL):
⚠️ Values with 3 digits after comma = THOUSANDS format (remove comma):
  - "1,880" → 1880 (one thousand eight hundred eighty)
  - "0,150" → 150 (one hundred fifty)
  - "2,000" → 2000 (two thousand)
  - "125,000" → 125000 (one hundred twenty-five thousand)
⚠️ Values with 1-2 digits after comma = DECIMAL format (replace comma with dot):
  - "1,88" → 1.88 (one point eighty-eight)
  - "2,5" → 2.5 (two point five)
  - "34,00" → 34.0 (thirty-four point zero)

Rule: If there are exactly 3 digits after the comma, it's thousands format. If 1-2 digits, it's decimal format.

EXAMPLES (quantity extraction):
✅ Correct Elastron (ignore volume):
- "34,00 ML 1 3,99" → quantidade=34.0 (ignore "1" = volume)
- "104,00 ML 2 1,98" → quantidade=104.0 (ignore "2" = volume)
- "48,50 ML 1 2,49" → quantidade=48.5 (ignore "1" = volume)

✅ Correct Eurospuma (ignore dimensions):
- "125,000 UN 1,880 0,150 0,080" → quantidade=125.0 (NOT 1.880)
- "640,000 MT 320,000 0,260" → quantidade=640.0 (NOT 320.0)
- "300,000 MT 30,000 2,100" → quantidade=300.0 (NOT 30.0)

✅ Correct normal:
- "COLCHAO | 2 UN | 199€" → quantidade=2
- "MATELAS | Qté: 5 | 245€" → quantidade=5

❌ Wrong:
- "34,00 ML 1" → quantidade=1 ✗ (should be 34.0)
- "125,000 UN 1,880" → quantidade=1.880 ✗ (should be 125.0)

Rules:
- Extract EVERY product line
- INTEGER ≤3 after unit = volume (Elastron only) → ignore
- DECIMAL after unit = dimension → don't use as quantity
- First number before unit = quantity (always)
- IMPORTANT: If document has multiple PO numbers (Encomenda nr, Pedido nr, etc), extract the PO number for EACH product"""

        user_prompt = f"""Extract ALL products from this document (PT/ES/FR):

{ocr_text[:3000] if ocr_text else "No OCR text"}

Return complete JSON with ALL products."""

        response = requests.post(
            "https://api.groq.com/openai/v1/chat/completions",
            headers={
                "Authorization": f"Bearer {api_key}",
                "Content-Type": "application/json"
            },
            json={
                "model": "llama-3.3-70b-versatile",
                "messages": [
                    {"role": "system", "content": system_prompt},
                    {"role": "user", "content": user_prompt}
                ],
                "temperature": 0.1,
                "max_tokens": 4000,
                "response_format": {"type": "json_object"}
            },
            timeout=30
        )

        if response.status_code == 200:
            result = response.json()
            content = result['choices'][0]['message']['content']
            
            extracted_data = json.loads(content)
            produtos_count = len(extracted_data.get('produtos', []))
            print(f"✅ Groq LLM ({key_name}): {produtos_count} produtos extraídos")
            return extracted_data, 200
        else:
            print(f"⚠️ Groq HTTP {response.status_code} ({key_name})")
            return None, response.status_code
            
    except Exception as e:
        print(f"⚠️ Groq exception ({key_name}): {e}")
        return None, 500

def ollama_extract_document(file_path: str, ocr_text: str = None):
    """
    LLM Document Extractor - Level -1 (pós-processador inteligente)
    
    Usa LLM (Groq/Ollama) para extrair dados estruturados de documentos.
    Combina texto OCR com prompt engineering.
    
    Args:
        file_path: Caminho do PDF/imagem
        ocr_text: Texto já extraído por OCR (opcional, melhora resultados)
    
    Returns:
        dict com metadados + produtos ou None se falhar
    """
    # Verificar se requests está disponível
    if not OCR_SPACE_AVAILABLE:
        print("⚠️ requests não disponível - LLM desabilitado")
        return None
    
    # Tentar Groq primeiro com chave primária
    groq_key = os.environ.get('GROQ_API_KEY')
    if groq_key:
        print("🔑 Tentando Groq com chave primária (GROQ_API_KEY)...")
        groq_result, status_code = groq_extract_document(file_path, ocr_text, groq_key, "GROQ_API_KEY")
        
        # Se sucesso, retornar resultado
        if groq_result and groq_result.get('produtos'):
            return groq_result
        
        # Se erro 429 (rate limit), tentar segunda chave
        if status_code == 429:
            groq_key_2 = os.environ.get('GROQ_API_KEY_2')
            if groq_key_2:
                print("🔄 Rate limit na chave primária - tentando chave secundária (GROQ_API_KEY_2)...")
                groq_result_2, status_code_2 = groq_extract_document(file_path, ocr_text, groq_key_2, "GROQ_API_KEY_2")
                
                if groq_result_2 and groq_result_2.get('produtos'):
                    return groq_result_2
                
                print(f"⚠️ Groq chave secundária também falhou (status {status_code_2})")
            else:
                print("⚠️ GROQ_API_KEY_2 não configurada - sem fallback disponível")
        
        print("⚠️ Groq falhou ou sem produtos - tentando Ollama fallback")
    
    # Fallback: Ollama (se configurado)
    ollama_url = os.environ.get('OLLAMA_API_URL')
    ollama_model = os.environ.get('OLLAMA_MODEL', 'llama3.2-vision')
    
    if not ollama_url:
        print("⚠️ Nenhum LLM configurado ou todos falharam")
        return None
    
    try:
        # Prompt melhorado com exemplos concretos PT/ES/FR
        system_prompt = """You are a document extraction expert. Extract ALL product data from invoices, delivery notes, and purchase orders in Portuguese, Spanish, or French.

CRITICAL: Extract EVERY product line, even if incomplete or malformed.

ALWAYS return valid JSON with this exact structure:
{
  "fornecedor": "supplier name or null",
  "nif": "tax ID or null",
  "numero_documento": "document number or null",
  "data_documento": "YYYY-MM-DD or null",
  "numero_encomenda": "PO number or null",
  "iban": "bank account or null",
  "produtos": [
    {
      "codigo": "product code/SKU",
      "descricao": "product description",
      "quantidade": 10.5,
      "preco_unitario": 25.99,
      "total": 272.40,
      "numero_encomenda": "PO number for this specific product (if document has multiple POs)"
    }
  ],
  "total_geral": 272.40
}

NUMBER FORMATTING (CRITICAL):
⚠️ Values with 3 digits after comma = THOUSANDS format (remove comma):
  - "1,880" → 1880 (one thousand eight hundred eighty)
  - "0,150" → 150 (one hundred fifty)
  - "2,000" → 2000 (two thousand)
  - "125,000" → 125000 (one hundred twenty-five thousand)
⚠️ Values with 1-2 digits after comma = DECIMAL format (replace comma with dot):
  - "1,88" → 1.88 (one point eighty-eight)
  - "2,5" → 2.5 (two point five)
  - "34,00" → 34.0 (thirty-four point zero)

Rule: If there are exactly 3 digits after the comma, it's thousands format. If 1-2 digits, it's decimal format.

EXAMPLES:

Portuguese document:
COLCHAO VISCO 150X190 | 2 UN | 199.00€ | 398.00€
→ {"codigo": "VISCO150", "descricao": "COLCHAO VISCO 150X190", "quantidade": 2, "preco_unitario": 199.00, "total": 398.00}

Spanish document (multi-line):
4,00
COLCHON TOP VISCO 2019 135X190
LUSTOPVS135190
→ {"codigo": "LUSTOPVS135190", "descricao": "COLCHON TOP VISCO 2019 135X190", "quantidade": 4.0, "preco_unitario": null, "total": null}

French document:
MATELAS SAN REMO 140X200 | Qté: 2 | PU: 245,00€ | Total: 490,00€
→ {"codigo": "SANREMO140", "descricao": "MATELAS SAN REMO 140X200", "quantidade": 2, "preco_unitario": 245.00, "total": 490.00}

Document with MULTIPLE Purchase Orders (extract PO number per product):
Encomenda nr 11-161050
PRODUTO-A | 10 UN
Encomenda nr 11-161594  
PRODUTO-B | 5 UN
PRODUTO-C | 8 UN
→ [
  {"codigo": "PRODUTO-A", "quantidade": 10, "numero_encomenda": "11-161050"},
  {"codigo": "PRODUTO-B", "quantidade": 5, "numero_encomenda": "11-161594"},
  {"codigo": "PRODUTO-C", "quantidade": 8, "numero_encomenda": "11-161594"}
]

Rules:
- Extract EVERY product, even if some fields are missing
- Products can be in tables, lists, or multi-line format
- Ignore addresses, headers, footers - ONLY extract products
- Convert all quantities/prices to numbers (replace comma with dot)
- Use null for missing fields
- IMPORTANT: If document has multiple PO numbers (Encomenda nr, Pedido nr, etc), extract the PO number for EACH product
- Return ONLY the JSON, no markdown, no explanations"""

        # Preparar preview do texto OCR (sem backslash em f-string)
        ocr_preview = f"OCR Text:\n{ocr_text[:2000]}" if ocr_text else "No OCR text - analyze image directly"
        
        user_prompt = f"""Extract ALL product data from this document.

Document type: Invoice/Delivery Note/Purchase Order (PT/ES/FR)

{ocr_preview}

IMPORTANT:
- Extract EVERY product, even if data is incomplete or spread across multiple lines
- Look for patterns: quantities + descriptions + codes
- Ignore addresses, legal text, headers/footers
- Return complete JSON with ALL products found"""

        # Preparar request para Ollama
        payload = {
            "model": ollama_model,
            "messages": [
                {"role": "system", "content": system_prompt},
                {"role": "user", "content": user_prompt}
            ],
            "stream": False,
            "format": "json",  # Force JSON output
            "options": {
                "temperature": 0.1,  # Baixa criatividade para dados estruturados
                "num_predict": 4000,  # Tokens para documentos muito grandes
                "top_p": 0.9
            }
        }
        
        # Se modelo suporta vision, adicionar imagem
        if 'vision' in ollama_model.lower() and file_path.lower().endswith('.pdf'):
            # Converter primeira página PDF para base64
            try:
                images = convert_from_path(file_path, first_page=1, last_page=1, dpi=150)
                if images:
                    img_buffer = BytesIO()
                    images[0].save(img_buffer, format='PNG')
                    img_base64 = base64.b64encode(img_buffer.getvalue()).decode('utf-8')
                    
                    payload["messages"][-1]["images"] = [img_base64]
                    print(f"✅ Ollama vision: imagem adicionada ({len(img_base64)} bytes)")
            except Exception as e:
                print(f"⚠️ Erro ao converter PDF para imagem: {e}")
        
        # Chamar Ollama API
        print(f"🤖 Ollama ({ollama_model}): processando documento...")
        print(f"   URL: {ollama_url}/api/chat")
        print(f"   OCR context: {len(ocr_text) if ocr_text else 0} chars")
        print(f"   Timeout: 60s")
        
        response = requests.post(
            f"{ollama_url}/api/chat",
            json=payload,
            timeout=60  # 60s timeout para LLMs (mais lento que OCR)
        )
        
        if response.status_code == 200:
            result = response.json()
            content = result.get('message', {}).get('content', '')
            
            if not content:
                print("⚠️ Ollama retornou resposta vazia")
                return None
            
            # Parse JSON da resposta
            try:
                # Limpar markdown code blocks se existirem
                if '```json' in content:
                    content = content.split('```json')[1].split('```')[0].strip()
                elif '```' in content:
                    content = content.split('```')[1].split('```')[0].strip()
                
                extracted_data = json.loads(content)
                
                # Validar estrutura mínima
                if not isinstance(extracted_data, dict):
                    print("⚠️ Ollama retornou JSON inválido (não é dict)")
                    return None
                
                produtos_count = len(extracted_data.get('produtos', []))
                print(f"✅ Ollama: {produtos_count} produtos extraídos")
                print(f"   Fornecedor: {extracted_data.get('fornecedor', 'N/A')}")
                print(f"   Doc: {extracted_data.get('numero_documento', 'N/A')}")
                
                return extracted_data
                
            except json.JSONDecodeError as e:
                print(f"⚠️ Ollama retornou JSON inválido: {e}")
                print(f"   Resposta: {content[:200]}...")
                return None
        else:
            print(f"⚠️ Ollama HTTP {response.status_code}: {response.text[:200]}")
            return None
            
    except requests.Timeout:
        print("⚠️ Ollama timeout (60s) - fallback para OCR")
        return None
    except Exception as e:
        print(f"⚠️ Ollama exception: {e}")
        return None

# --- PaddleOCR (lazy loading para evitar problemas no startup) ---
_paddle_ocr_instance = None

def get_paddle_ocr():
    """Inicializa PaddleOCR lazy - só quando necessário."""
    global _paddle_ocr_instance
    if _paddle_ocr_instance is None:
        try:
            from paddleocr import PaddleOCR
            _paddle_ocr_instance = PaddleOCR(use_angle_cls=True, lang='pt')
            print("✅ PaddleOCR inicializado (português)")
        except Exception as e:
            print(f"⚠️ PaddleOCR não disponível: {e}")
            _paddle_ocr_instance = False
    return _paddle_ocr_instance if _paddle_ocr_instance is not False else None

# --- EasyOCR (lazy loading para evitar problemas no startup) ---
_easyocr_instance = None

def get_easy_ocr():
    """Inicializa EasyOCR lazy - só quando necessário."""
    global _easyocr_instance
    if _easyocr_instance is None:
        try:
            import easyocr
            _easyocr_instance = easyocr.Reader(['pt', 'es', 'fr'], gpu=False)
            print("✅ EasyOCR inicializado (PT/ES/FR)")
        except Exception as e:
            print(f"⚠️ EasyOCR não disponível: {e}")
            _easyocr_instance = False
    return _easyocr_instance if _easyocr_instance is not False else None

# ----------------- OCR: PDF/Imagens -----------------


def save_extraction_to_json(data: dict, filename: str = "extracao.json"):
    """Salva os dados extraídos em um arquivo JSON."""
    try:
        json_path = os.path.join(settings.BASE_DIR, filename)
        with open(json_path, 'w', encoding='utf-8') as f:
            json.dump(data, f, ensure_ascii=False, indent=2)
        print(f"✅ Dados salvos em {json_path}")
        return json_path
    except Exception as e:
        print(f"❌ Erro ao salvar JSON: {e}")
        return None


def real_ocr_extract(file_path: str):
    """OCR usando Tesseract. Extrai texto e faz parse para estrutura."""
    text_content = ""
    qr_codes = []
    ext = os.path.splitext(file_path)[1].lower()

    print(f"🔍 Processando com Tesseract: {os.path.basename(file_path)}")
    
    # Flag para controlar se usamos PyMuPDF
    used_pymupdf = False
    
    if ext == ".pdf":
        text_content, qr_codes = extract_text_from_pdf(file_path)
        # Verificar se PyMuPDF foi usado (texto contém marcadores de página)
        if "--- Página" in text_content:
            used_pymupdf = True
    elif ext in [".jpg", ".jpeg", ".png", ".tiff", ".bmp"]:
        text_content, qr_codes = extract_text_from_image(file_path)

    # Validação antecipada: se texto muito curto, pode ser ficheiro ilegível/desformatado
    texto_pdfplumber_curto = len(text_content) < 50
    if texto_pdfplumber_curto:
        print(f"⚠️ Texto pdfplumber muito curto ({len(text_content)} chars) - possível ficheiro ilegível")

    preview = "\n".join(text_content.splitlines()[:60])
    print("---- OCR PREVIEW (primeiras linhas) ----")
    print(preview)
    print("----------------------------------------")

    if qr_codes:
        print(f"✅ {len(qr_codes)} QR code(s) detectado(s)")

    if not text_content.strip():
        print("❌ OCR vazio")
        error_result = {
            "error": "OCR failed - no text extracted from document",
            "numero_requisicao": f"ERROR-{os.path.basename(file_path)}",
            "document_number": "",
            "po_number": "",
            "supplier_name": "",
            "delivery_date": "",
            "qr_codes": qr_codes,
            "lines": [],
            "totals": {
                "total_lines": 0,
                "total_quantity": 0
            },
        }
        save_extraction_to_json(error_result)
        return error_result

    result = parse_portuguese_document(text_content, qr_codes, texto_pdfplumber_curto, file_path=file_path)
    
    # FALLBACK INTELIGENTE: Se PyMuPDF foi usado mas nenhum produto foi extraído,
    # fazer fallback para pdfplumber que tem melhor compatibilidade com parsers existentes
    produtos_extraidos = len(result.get('produtos', []))
    if used_pymupdf and produtos_extraidos == 0 and ext == ".pdf":
        # Verificar se o texto contém códigos de produtos (indica que há dados para extrair)
        import re
        tem_codigos = bool(re.search(r'\b[A-Z]{3}\d{12}\b', text_content))
        
        if tem_codigos:
            print("🔄 PyMuPDF extraiu texto mas parser retornou 0 produtos - fallback para pdfplumber...")
            # Forçar uso de pdfplumber/Tesseract
            try:
                import pdfplumber
                text_pdfplumber = ""
                with pdfplumber.open(file_path) as pdf:
                    for page in pdf.pages:
                        page_text = page.extract_text() or ""
                        text_pdfplumber += page_text + "\n"
                
                if text_pdfplumber.strip():
                    print(f"✅ pdfplumber fallback: {len(text_pdfplumber)} chars")
                    result = parse_portuguese_document(text_pdfplumber, qr_codes, False, file_path=file_path)
                    produtos_extraidos = len(result.get('produtos', []))
                    print(f"✅ Fallback result: {produtos_extraidos} produtos extraídos")
            except Exception as e:
                print(f"⚠️ Fallback pdfplumber falhou: {e}")
    
    save_extraction_to_json(result)
    return result


def extract_with_pymupdf(file_path: str):
    """
    Extrai texto de PDF usando PyMuPDF (fitz) preservando layout e estrutura.
    Melhor que PyPDF2 para tabelas multi-página e preservação de layout.
    
    Returns:
        tuple: (texto_extraido, qr_codes) ou None se falhar
    """
    if not PYMUPDF_AVAILABLE:
        return None
    
    try:
        doc = fitz.open(file_path)
        text_parts = []
        
        print(f"📄 PyMuPDF: Processando {len(doc)} página(s)...")
        
        for page_num, page in enumerate(doc, start=1):
            # Extrai texto mantendo layout original (melhor para tabelas)
            page_text = page.get_text("text")
            
            if page_text.strip():
                text_parts.append(f"--- Página {page_num} ---\n{page_text}\n")
        
        doc.close()
        
        full_text = "\n".join(text_parts)
        
        if full_text.strip() and len(full_text.strip()) > 50:
            print(f"✅ PyMuPDF extraction: {len(full_text)} chars de {len(text_parts)} página(s)")
            
            # Detectar QR codes
            qr_codes = []
            if QR_CODE_ENABLED:
                try:
                    print("🔍 Procurando QR codes no PDF...")
                    pages = convert_from_path(file_path, dpi=300)
                    for page_num, page_img in enumerate(pages, start=1):
                        page_qr = detect_and_read_qrcodes(page_img, page_number=page_num)
                        qr_codes.extend(page_qr)
                except Exception as e:
                    print(f"⚠️ Erro ao buscar QR codes: {e}")
            
            return full_text.strip(), qr_codes
        
        return None
        
    except Exception as e:
        print(f"⚠️ PyMuPDF falhou: {e}")
        return None


def extract_text_from_pdf(file_path: str):
    """
    Cascata de extração de PDF (5 níveis):
    1. PyMuPDF (fitz) - melhor layout e multi-página
    2. PyPDF2 texto embutido - mais rápido como fallback
    3. OCR.space API - cloud, preciso, grátis 25k/mês
    4. PaddleOCR/EasyOCR/Tesseract - local, offline
    """
    try:
        # LEVEL 1: PyMuPDF (melhor para layout e multi-página)
        pymupdf_result = extract_with_pymupdf(file_path)
        if pymupdf_result:
            return pymupdf_result
        
        # LEVEL 2: Tenta texto embutido com PyPDF2 (fallback)
        text = ""
        with open(file_path, "rb") as f:
            reader = PyPDF2.PdfReader(f)
            for page in reader.pages:
                page_text = page.extract_text() or ""
                text += page_text + "\n"

        if text.strip() and len(text.strip()) > 50:
            print(f"✅ PyPDF2 text extraction: {len(text)} chars")
            # Mesmo com texto embutido, tenta detectar QR codes
            qr_codes = []
            if QR_CODE_ENABLED:
                try:
                    print("🔍 Procurando QR codes no PDF...")
                    pages = convert_from_path(file_path, dpi=300)
                    for page_num, page_img in enumerate(pages, start=1):
                        page_qr = detect_and_read_qrcodes(page_img,
                                                          page_number=page_num)
                        qr_codes.extend(page_qr)
                except Exception as e:
                    print(f"⚠️ Erro ao buscar QR codes: {e}")
            return text.strip(), qr_codes

        # LEVEL 3: OCR.space API (cloud, grátis, preciso)
        print("📄 PDF sem texto embutido - tentando OCR.space API...")
        ocr_text = ocr_space_api(file_path, language='por')
        
        if ocr_text and len(ocr_text.strip()) > 50:
            # QR codes (se disponível)
            qr_codes = []
            if QR_CODE_ENABLED:
                try:
                    print("🔍 Procurando QR codes no PDF...")
                    pages = convert_from_path(file_path, dpi=300)
                    for page_num, page_img in enumerate(pages, start=1):
                        page_qr = detect_and_read_qrcodes(page_img, page_number=page_num)
                        qr_codes.extend(page_qr)
                except Exception as e:
                    print(f"⚠️ Erro ao buscar QR codes: {e}")
            return ocr_text.strip(), qr_codes

        # LEVEL 4: Engines locais (PaddleOCR → EasyOCR → Tesseract)
        print("📄 OCR.space falhou - usando engines locais (PaddleOCR/EasyOCR/Tesseract)...")
        return extract_text_from_pdf_with_ocr(file_path)

    except Exception as e:
        print(f"❌ Erro no extract_text_from_pdf: {e}")
        return extract_text_from_pdf_with_ocr(file_path)


def parse_qrcode_fiscal_pt(qr_data: str):
    """Parse de QR code fiscal português (formato A:valor*B:valor*...) com nomes descritivos."""

    # Mapeamento dos códigos para nomes descritivos (Especificações Técnicas AT)
    FIELD_NAMES = {
        "A": "nif_emitente",
        "B": "nif_adquirente",
        "C": "pais_adquirente",
        "D": "tipo_documento",
        "E": "estado_documento",
        "F": "data_documento",
        "G": "identificacao_documento",
        "H": "atcud",
        "I1": "espaco_fiscal",
        "I2": "base_tributavel_isenta_iva",
        "I3": "base_tributavel_taxa_reduzida",
        "I4": "total_iva_taxa_reduzida",
        "I5": "base_tributavel_taxa_intermedia",
        "I6": "total_iva_taxa_intermedia",
        "I7": "base_tributavel_taxa_normal",
        "I8": "total_iva_taxa_normal",
        "J1": "espaco_fiscal_2",
        "J2": "base_tributavel_isenta_iva_2",
        "J3": "base_tributavel_taxa_reduzida_2",
        "J4": "total_iva_taxa_reduzida_2",
        "J5": "base_tributavel_taxa_intermedia_2",
        "J6": "total_iva_taxa_intermedia_2",
        "J7": "base_tributavel_taxa_normal_2",
        "J8": "total_iva_taxa_normal_2",
        "K1": "espaco_fiscal_3",
        "K2": "base_tributavel_isenta_iva_3",
        "K3": "base_tributavel_taxa_reduzida_3",
        "K4": "total_iva_taxa_reduzida_3",
        "K5": "base_tributavel_taxa_intermedia_3",
        "K6": "total_iva_taxa_intermedia_3",
        "K7": "base_tributavel_taxa_normal_3",
        "K8": "total_iva_taxa_normal_3",
        "L": "nao_sujeito_iva",
        "M": "imposto_selo",
        "N": "total_impostos",
        "O": "total_documento",
        "P": "retencao_na_fonte",
        "Q": "hash",
        "R": "certificado",
        "S": "outras_infos"
    }

    try:
        if not qr_data or "*" not in qr_data:
            return None

        parsed_raw = {}
        fields = qr_data.split("*")

        for field in fields:
            if ":" in field:
                key, value = field.split(":", 1)
                parsed_raw[key] = value

        # Valida se é realmente um QR fiscal português
        # QR fiscal deve ter pelo menos o campo A (NIF emitente)
        if not parsed_raw or "A" not in parsed_raw:
            return None

        # Converte para nomes descritivos
        parsed = {}
        for code, value in parsed_raw.items():
            field_name = FIELD_NAMES.get(code, code)
            parsed[field_name] = value

        return parsed if parsed else None
    except Exception as e:
        print(f"⚠️ Erro ao parsear QR fiscal: {e}")
        return None


def detect_and_read_qrcodes(image, page_number=None):
    """Lê QR codes usando OpenCV e retorna lista estruturada."""
    if not QR_CODE_ENABLED:
        return []

    try:
        arr = np.array(image)
        if len(arr.shape) == 3 and arr.shape[2] == 3:
            arr = cv2.cvtColor(arr, cv2.COLOR_RGB2BGR)
        elif len(arr.shape) == 3 and arr.shape[2] == 4:
            arr = cv2.cvtColor(arr, cv2.COLOR_RGBA2BGR)

        # Usa o detector de QR code do OpenCV
        detector = cv2.QRCodeDetector()
        data, vertices_array, _ = detector.detectAndDecode(arr)

        result = []
        if vertices_array is not None and data:
            print(f"✅ QR: {data[:80]}…")

            # Tenta parsear QR code fiscal português
            parsed = parse_qrcode_fiscal_pt(data)
            if parsed:
                # Se parseou com sucesso, coloca os dados estruturados no campo "data"
                qr_info = {"data": parsed, "raw_data": data}
            else:
                # Se não conseguiu parsear, mantém como string
                qr_info = {"data": data}

            if page_number is not None:
                qr_info["page"] = page_number

            result.append(qr_info)

        # Tenta detectar múltiplos QR codes (OpenCV 4.5.4+)
        try:
            multi_data = detector.detectAndDecodeMulti(arr)
            if multi_data[0]:  # Se detectou algum
                for i, qr_data in enumerate(multi_data[1]):
                    # Verifica se já não foi adicionado
                    already_added = False
                    for existing in result:
                        if existing.get("raw_data") == qr_data or existing.get(
                                "data") == qr_data:
                            already_added = True
                            break

                    if qr_data and not already_added:
                        print(f"✅ QR: {qr_data[:80]}…")

                        # Tenta parsear QR code fiscal português
                        parsed = parse_qrcode_fiscal_pt(qr_data)
                        if parsed:
                            qr_info = {"data": parsed, "raw_data": qr_data}
                        else:
                            qr_info = {"data": qr_data}

                        if page_number is not None:
                            qr_info["page"] = page_number

                        result.append(qr_info)
        except:
            pass  # Versão do OpenCV pode não suportar detectAndDecodeMulti

        return result
    except Exception as e:
        print(f"⚠️ QR erro: {e}")
        return []


def extract_text_from_pdf_with_ocr(file_path: str):
    """Converte todas as páginas para imagem e aplica PaddleOCR (ou Tesseract como fallback)."""
    import time
    import numpy as np
    try:
        # Tenta usar PaddleOCR primeiro
        paddle_ocr = get_paddle_ocr()
        ocr_engine = "PaddleOCR" if paddle_ocr else "Tesseract"
        
        print(f"📄 Converter PDF → imagens (OCR com {ocr_engine})…")
        
        # Converter PDF → imagens primeiro
        start_time = time.time()
        pages = convert_from_path(file_path, dpi=300)  # DPI 300 para melhor qualidade
        conversion_time = time.time() - start_time
        
        # Se conversão demorou muito (>20s), ficheiro pode ter problemas
        if conversion_time > 20:
            print(f"⚠️ Conversão PDF demorou {conversion_time:.1f}s - possível ficheiro problemático")
        
        all_text = ""
        all_qr_codes = []
        
        for i, page in enumerate(pages, 1):
            print(f"🔍 Página {i}/{len(pages)} - {ocr_engine}")
            
            # Limite de tempo por página: 15 segundos
            page_start = time.time()
            
            qr_codes = detect_and_read_qrcodes(page, page_number=i)
            all_qr_codes.extend(qr_codes)
            
            # OCR da página - cascata de 3 níveis
            page_text = ""
            paddle_failed = False
            easy_failed = False
            ocr_engine_used = None
            
            try:
                # Nível 1: PaddleOCR (rápido e preciso)
                if paddle_ocr:
                    try:
                        img_array = np.array(page)
                        result = paddle_ocr.ocr(img_array, cls=True)
                        
                        if result and result[0]:
                            for line in result[0]:
                                if line and len(line) >= 2:
                                    text = line[1][0]
                                    confidence = line[1][1]
                                    if confidence > 0.5:
                                        page_text += text + "\n"
                        
                        if page_text.strip():
                            ocr_engine_used = "PaddleOCR"
                        else:
                            paddle_failed = True
                            print(f"⚠️ PaddleOCR não extraiu texto da página {i}, tentando EasyOCR...")
                    except Exception as paddle_error:
                        paddle_failed = True
                        print(f"⚠️ PaddleOCR falhou na página {i}: {paddle_error}, tentando EasyOCR...")
                
                # Nível 2: EasyOCR (se PaddleOCR falhou)
                if (not paddle_ocr or paddle_failed) and not page_text.strip():
                    easy_ocr = get_easy_ocr()
                    if easy_ocr:
                        try:
                            import numpy as np
                            img_array = np.array(page)
                            result = easy_ocr.readtext(img_array)
                            
                            if result:
                                for detection in result:
                                    text = detection[1]
                                    confidence = detection[2]
                                    if confidence > 0.3:
                                        page_text += text + " "
                                page_text = page_text.strip() + "\n"
                            
                            if page_text.strip():
                                ocr_engine_used = "EasyOCR"
                            else:
                                easy_failed = True
                                print(f"⚠️ EasyOCR não extraiu texto da página {i}, tentando Tesseract...")
                        except Exception as easy_error:
                            easy_failed = True
                            print(f"⚠️ EasyOCR falhou na página {i}: {easy_error}, tentando Tesseract...")
                
                # Nível 3: Tesseract (fallback final)
                if not page_text.strip():
                    page_text = pytesseract.image_to_string(
                        page, config="--psm 3 --oem 3 -l por", lang="por", timeout=60)
                    if page_text.strip():
                        ocr_engine_used = "Tesseract"
                
                if page_text.strip():
                    all_text += f"\n--- Página {i} ---\n{page_text}\n"
                    if ocr_engine_used:
                        print(f"✅ Página {i} processada com {ocr_engine_used}")
                    
            except RuntimeError as e:
                if "timeout" in str(e).lower():
                    print(f"⚠️ Timeout OCR na página {i} - imagem de má qualidade")
                else:
                    raise
            except Exception as e:
                print(f"⚠️ Erro OCR na página {i}: {e}")
            
            page_time = time.time() - page_start
            if page_time > 10:
                print(f"⚠️ Página {i} demorou {page_time:.1f}s - qualidade baixa")
        
        print(f"✅ OCR completo: {len(pages)} páginas")
        return all_text.strip(), all_qr_codes
    except Exception as e:
        print(f"❌ OCR PDF erro: {e}")
        return "", []


def extract_text_from_image(file_path: str):
    """OCR para imagem com cascata de 3 níveis: PaddleOCR → EasyOCR → Tesseract."""
    import numpy as np
    try:
        img = Image.open(file_path)
        qr_codes = detect_and_read_qrcodes(img)
        
        ocr_text = ""
        paddle_failed = False
        easy_failed = False
        ocr_engine_used = None
        
        # Nível 1: PaddleOCR (rápido e preciso)
        paddle_ocr = get_paddle_ocr()
        if paddle_ocr:
            try:
                img_array = np.array(img)
                result = paddle_ocr.ocr(img_array, cls=True)
                
                if result and result[0]:
                    for line in result[0]:
                        if line and len(line) >= 2:
                            text = line[1][0]
                            confidence = line[1][1]
                            if confidence > 0.5:
                                ocr_text += text + "\n"
                
                if ocr_text.strip():
                    ocr_engine_used = "PaddleOCR"
                else:
                    paddle_failed = True
                    print(f"⚠️ PaddleOCR não extraiu texto da imagem, tentando EasyOCR...")
            except Exception as paddle_error:
                paddle_failed = True
                print(f"⚠️ PaddleOCR falhou: {paddle_error}, tentando EasyOCR...")
        
        # Nível 2: EasyOCR (se PaddleOCR falhou)
        if (not paddle_ocr or paddle_failed) and not ocr_text.strip():
            easy_ocr = get_easy_ocr()
            if easy_ocr:
                try:
                    img_array = np.array(img)
                    result = easy_ocr.readtext(img_array)
                    
                    if result:
                        for detection in result:
                            text = detection[1]
                            confidence = detection[2]
                            if confidence > 0.3:
                                ocr_text += text + " "
                        ocr_text = ocr_text.strip() + "\n"
                    
                    if ocr_text.strip():
                        ocr_engine_used = "EasyOCR"
                    else:
                        easy_failed = True
                        print(f"⚠️ EasyOCR não extraiu texto da imagem, tentando Tesseract...")
                except Exception as easy_error:
                    easy_failed = True
                    print(f"⚠️ EasyOCR falhou: {easy_error}, tentando Tesseract...")
        
        # Nível 3: Tesseract (fallback final)
        if not ocr_text.strip():
            ocr_text = pytesseract.image_to_string(img,
                                                   config="--psm 6 --oem 3 -l por",
                                                   lang="por")
            if ocr_text.strip():
                ocr_engine_used = "Tesseract"
        
        if ocr_engine_used:
            print(f"✅ Imagem processada com {ocr_engine_used}")
        
        return ocr_text.strip(), qr_codes
    except Exception as e:
        print(f"❌ OCR imagem erro: {e}")
        return "", []


# ----------------- PARSE: heurísticas PT -----------------


def extract_guia_remessa_products(text: str):
    """
    Extrai produtos da tabela de Guia de Remessa com parser flexível.
    Campos: Artigo, Descrição, Lote Produção, Quant., Un., Vol., Preço Un., Desconto, Iva, Total
    """
    products = []
    lines = text.split("\n")

    current_ref = ""

    # Regex para detectar referências de ordem (mais flexível)
    ref_pattern = re.compile(
        r"^\s*(\d[A-Z]{2,6}\s+N[oº°]\s*\d+[/\-]\d+[A-Z]{0,4}\s+de\s+\d{2}-\d{2}-\d{4})",
        re.IGNORECASE)

    # Regex mais flexível para linha de produto
    # Formato: E0748001901  131,59 1  34,00 3,00 ML 3,99 23,00 5159-250602064 BALTIC fb, TOFFEE
    # Artigo: letras + números (mais flexível)
    # Volume: pode ser decimal
    # Lote: pode estar vazio ou ter vários formatos
    # Unidade: pode ter 2-10 caracteres
    product_pattern = re.compile(
        r"^([A-Z]+\d+[A-Z0-9]*)\s+"  # Artigo (flexível: E0748001901, ABC123, etc.)
        r"([\d,\.]+)\s+"  # Total
        r"([\d,\.]+)\s+"  # Volume (aceita decimais)
        r"([\d,\.]+)\s+"  # Quantidade
        r"([\d,\.]+)\s+"  # Desconto
        r"([A-Z]{2,10})\s+"  # Unidade (mais flexível)
        r"([\d,\.]+)\s+"  # Preço Unitário
        r"([\d,\.]+)\s+"  # IVA
        r"([\w\-#]*)\s*"  # Lote (opcional, pode estar vazio)
        r"(.+?)\s*$",  # Descrição (resto da linha)
        re.IGNORECASE)

    for line in lines:
        stripped = line.strip()

        # Verifica se é uma referência de ordem
        ref_match = ref_pattern.match(stripped)
        if ref_match:
            current_ref = ref_match.group(1).strip()
            continue

        # Verifica se é uma linha de produto
        prod_match = product_pattern.match(stripped)
        if prod_match:
            try:
                artigo = prod_match.group(1).strip()
                total = normalize_number(prod_match.group(2))
                volume = normalize_number(prod_match.group(3))
                quantidade = normalize_number(prod_match.group(4))
                desconto = normalize_number(prod_match.group(5))
                unidade = prod_match.group(6).strip()
                preco_un = normalize_number(prod_match.group(7))
                iva = normalize_number(prod_match.group(8))
                lote = prod_match.group(9).strip() if prod_match.group(
                    9) else ""
                descricao = prod_match.group(10).strip()

                # Validações básicas
                if not artigo or not descricao:
                    continue

                product = {
                    "referencia_ordem": current_ref if current_ref else None,
                    "artigo": artigo,
                    "descricao": descricao,
                    "lote_producao": lote if lote else None,
                    "quantidade": quantidade,
                    "unidade": unidade,
                    "volume": volume,
                    "preco_unitario": preco_un,
                    "desconto": desconto,
                    "iva": iva,
                    "total": total
                }

                products.append(product)
            except (ValueError, IndexError) as e:
                print(
                    f"⚠️ Erro ao parsear linha de produto '{stripped[:50]}...': {e}"
                )
                continue

    if products:
        print(f"✅ Extraídos {len(products)} produtos da Guia de Remessa")
    else:
        print("⚠️ Nenhum produto encontrado no formato Guia de Remessa")

    return products


def detect_document_type(text: str):
    """Detecta automaticamente o tipo de documento português, francês e espanhol."""
    text_lower = text.lower()
    
    # Documentos espanhóis
    if ("pedido" in text_lower and ("españa" in text_lower or "spain" in text_lower)) or \
       ("pedido" in text_lower and any(kw in text_lower for kw in ["artículo", "articulo", "descripción", "descripcion", "unidades", "cantidad"])):
        return "PEDIDO_ESPANHOL"
    
    # Documentos franceses
    if "bon de commande" in text_lower or ("commande" in text_lower and "désignation" in text_lower):
        return "BON_COMMANDE"
    
    # Documentos portugueses
    if "ordem compra" in text_lower or "ordem de compra" in text_lower:
        return "ORDEM_COMPRA"
    elif "elastron" in text_lower and "fatura" in text_lower:
        return "FATURA_ELASTRON"
    elif "colmol" in text_lower and ("guia" in text_lower or "comunicação de saída" in text_lower):
        return "GUIA_COLMOL"
    elif "fatura" in text_lower or "ft" in text_lower:
        return "FATURA_GENERICA"
    elif "guia de remessa" in text_lower or "guia remessa" in text_lower:
        return "GUIA_GENERICA"
    elif "recibo" in text_lower or "receipt" in text_lower:
        return "RECIBO"
    else:
        return "DOCUMENTO_GENERICO"


def parse_fatura_elastron(text: str):
    """Parser específico para faturas Elastron (compatível com Tesseract)."""
    produtos = []
    lines = text.split("\n")
    
    current_ref = ""
    for i, line in enumerate(lines):
        line_stripped = line.strip()
        
        if re.match(r'^\d[A-Z]{4}\s+[NnºN]', line_stripped):
            current_ref = line_stripped
            continue
        
        artigo_match = re.match(r'^(E[O0]\d{9,10})\s+(.+)', line_stripped)
        if artigo_match:
            try:
                artigo = artigo_match.group(1).replace('O', '0')
                resto = artigo_match.group(2).strip()
                
                parts = resto.split()
                if len(parts) < 6:
                    continue
                
                # Tesseract format: TOTAL VOL QUANT DESC UNID PRECO IVA LOTE... DESCRICAO
                # Encontrar unidade (ML, MT, UN)
                unidade_idx = -1
                unidade = "ML"
                for idx, part in enumerate(parts):
                    if part.upper() in ['ML', 'MT', 'UN', 'M²', 'M2']:
                        unidade = part.upper()
                        unidade_idx = idx
                        break
                
                if unidade_idx < 3:
                    continue
                
                # Campos antes da unidade: TOTAL VOL QUANT DESC
                total = normalize_number(parts[0])
                volume = int(parts[1]) if parts[1].isdigit() else 1
                quantidade = normalize_number(parts[2])
                desconto = normalize_number(parts[3])
                
                # Campos depois da unidade: PRECO IVA LOTE ... DESCRICAO
                preco_un = normalize_number(parts[unidade_idx + 1]) if unidade_idx + 1 < len(parts) else 0.0
                iva = normalize_number(parts[unidade_idx + 2]) if unidade_idx + 2 < len(parts) else 23.0
                
                # Lote e descrição
                lote = ""
                descricao = ""
                if unidade_idx + 3 < len(parts):
                    remaining = ' '.join(parts[unidade_idx + 3:])
                    lote_match = re.search(r'(\d{4}-\d+(?:#)?)', remaining)
                    if lote_match:
                        lote = lote_match.group(1)
                        descricao = remaining[lote_match.end():].strip()
                    else:
                        descricao = remaining
                
                produtos.append({
                    "referencia_ordem": current_ref,
                    "artigo": artigo,
                    "descricao": descricao,
                    "lote_producao": lote,
                    "quantidade": quantidade,
                    "unidade": unidade,
                    "volume": volume,
                    "preco_unitario": preco_un,
                    "desconto": desconto,
                    "iva": iva,
                    "total": total
                })
            except (ValueError, IndexError) as e:
                print(f"⚠️ Erro ao parsear linha Elastron '{line_stripped[:60]}': {e}")
                continue
    
    return produtos


def parse_guia_colmol(text: str):
    """Parser específico para Guias de Remessa Colmol."""
    produtos = []
    lines = text.split("\n")
    
    current_encomenda = ""
    current_requisicao = ""
    
    for i, line in enumerate(lines):
        line_stripped = line.strip()
        
        if "ENCOMENDA Nº" in line_stripped:
            encomenda_match = re.search(r'ENCOMENDA Nº\.?\s*(\d+-\d+)', line_stripped)
            requisicao_match = re.search(r'REQUISICAO Nº\.?\s*(\d+)', line_stripped)
            if encomenda_match:
                current_encomenda = encomenda_match.group(1)
            if requisicao_match:
                current_requisicao = requisicao_match.group(1)
            continue
        
        if re.match(r'^[A-Z0-9]{10,}', line_stripped):
            parts = line_stripped.split()
            # Tolerância: Aceitar linhas com pelo menos 3 partes (código + descrição + algo)
            if len(parts) >= 3:
                try:
                    codigo = parts[0]
                    
                    descricao_parts = []
                    j = 1
                    # Parar na descrição quando encontrar: número decimal, unidade (UN/MT/ML), ou padrão CX.
                    while j < len(parts):
                        part = parts[j]
                        # Número decimal (quantidade)
                        if re.match(r'^\d+[.,]\d+$', part):
                            break
                        # Unidades conhecidas (às vezes vem antes da quantidade)
                        if part.upper() in ['UN', 'MT', 'ML', 'M²', 'M2']:
                            break
                        # Padrão de dimensões (CX.1150x...)
                        if re.match(r'^CX\.\d', part, re.IGNORECASE):
                            descricao_parts.append(part)
                            j += 1
                            break
                        descricao_parts.append(part)
                        j += 1
                    
                    descricao = ' '.join(descricao_parts)
                    
                    # Agora procurar quantidade (pode ter espaços antes)
                    while j < len(parts) and not re.match(r'^\d+[.,]\d+$', parts[j]):
                        j += 1
                    
                    # FALLBACK: Se não encontrar quantidade válida, tentar extrair de qualquer número na linha
                    quantidade = 0.0
                    if j < len(parts):
                        quantidade = normalize_number(parts[j])
                    else:
                        # Buscar qualquer número na linha (fallback para dados corrompidos)
                        for part in parts[1:]:  # Skip código (parts[0])
                            try:
                                num = normalize_number(part)
                                if num > 0:
                                    quantidade = num
                                    break
                            except:
                                continue
                    
                    unidade = parts[j+1] if j+1 < len(parts) else "UN"
                    med1 = normalize_number(parts[j+2]) if j+2 < len(parts) else 0.0
                    med2 = normalize_number(parts[j+3]) if j+3 < len(parts) else 0.0
                    med3 = normalize_number(parts[j+4]) if j+4 < len(parts) else 0.0
                    peso = normalize_number(parts[j+5]) if j+5 < len(parts) else 0.0
                    iva = normalize_number(parts[j+6]) if j+6 < len(parts) else 23.0
                    
                    # Só adicionar se tiver código E (descrição OU quantidade válida)
                    if codigo and (descricao or quantidade > 0):
                        produtos.append({
                            "referencia_ordem": f"{current_encomenda} / Req {current_requisicao}",
                            "artigo": codigo,
                            "descricao": descricao,
                            "lote_producao": "",
                            "quantidade": quantidade,
                            "unidade": unidade,
                            "volume": 0,
                            "dimensoes": f"{med1}x{med2}x{med3}",
                            "peso": peso,
                            "iva": iva,
                            "total": 0.0
                        })
                except (ValueError, IndexError) as e:
                    print(f"⚠️ Erro ao parsear linha Colmol '{line_stripped[:80]}': {e}")
                    continue
    
    return produtos


def parse_guia_generica(text: str):
    """
    Parser genérico para extrair produtos de qualquer formato de guia de remessa.
    Usa heurísticas para detectar tabelas com produtos.
    
    Suporta formatos complexos onde descrição contém números:
    - CBAGD00067 CX EUROSPUMA 3044 VE 125,000 UN 1,880 0,150 0,080 84,600 KG
    - Extrai quantidade correta (125,000) ignorando números na descrição (3044)
    """
    produtos = []
    lines = text.split("\n")
    
    pedido_atual = ""
    
    # Unidades conhecidas (ordenadas por especificidade)
    UNIDADES_CONHECIDAS = r'\b(UN|UNI|UNID|UNIDADES|MT|M2|M²|M3|M³|KG|G|ML|L|CX|PC|PCS|PAR|SET|RL|FD|PAC)\b'
    
    for i, line in enumerate(lines):
        stripped = line.strip()
        if not stripped or len(stripped) < 10:
            continue
        
        pedido_match = re.search(r'(?:PEDIDO|ORDER|ENCOMENDA)\s*[:/]?\s*(\d+)', stripped, re.IGNORECASE)
        if pedido_match:
            pedido_atual = pedido_match.group(1)
            continue
        
        # Estratégia 1: Procurar UNIDADE de quantidade (UN/MT) + pegar número IMEDIATAMENTE antes
        # Importante: Linhas podem ter múltiplas unidades (125,000 UN ... 84,600 KG)
        # Usar PRIMEIRA unidade de quantidade (não peso) e número adjacente
        # Exemplo: CBAGD00067 CX EUROSPUMA 3044 VE 125,000 UN 1,880 0,150 0,080 84,600 KG
        #          → quantidade=125,000 UN (não 84,600 KG que é peso)
        codigo_match = re.match(r'^([A-Z0-9]{8,})\s+(.+)', stripped, re.IGNORECASE)
        if codigo_match:
            codigo = codigo_match.group(1).strip()
            resto_linha = codigo_match.group(2).strip()
            
            # Procurar padrão: [NÚMERO] [ESPAÇO] [UNIDADE_QUANTIDADE]
            # Unidades de quantidade (não peso): UN, MT, M2, M², PC, CX, etc.
            # Excluir KG/G (peso) da primeira procura
            qtd_unidades = r'(UN|UNI|UNID|UNIDADES|MT|M2|M²|M3|M³|ML|L|CX|PC|PCS|PAR|SET|RL|FD|PAC)'
            
            # Procurar: [número] [espaço(s)] [UNIDADE]
            # Regex: ([\d,\.]+)\s+(UN|MT|etc.)
            qtd_pattern = rf'([\d,\.]+)\s+{qtd_unidades}\b'
            qtd_match = re.search(qtd_pattern, resto_linha, re.IGNORECASE)
            
            if qtd_match:
                quantidade_str = qtd_match.group(1).strip()
                unidade = qtd_match.group(2).upper()
                
                # Pegar a descrição: tudo ANTES do match de quantidade
                pos_qtd_inicio = qtd_match.start()
                descricao = resto_linha[:pos_qtd_inicio].strip()
                
                print(f"✅ Parser genérico Estratégia 1: {codigo} | {descricao} | {quantidade_str} {unidade}")
                
                try:
                    # Usar função de normalização (3 casas decimais = milhares)
                    quantidade = normalize_number(quantidade_str)
                    
                    dims = ""
                    dim_match = re.search(r'(\d{3,4})[xX×](\d{3,4})[xX×](\d{3,4})', descricao)
                    if dim_match:
                        dims = f"{float(dim_match.group(1))/1000:.2f}x{float(dim_match.group(2))/1000:.2f}x{float(dim_match.group(3))/1000:.2f}"
                    
                    produtos.append({
                        "referencia_ordem": pedido_atual or "",
                        "artigo": codigo,
                        "descricao": descricao,
                        "lote_producao": "",
                        "quantidade": quantidade,
                        "unidade": unidade,
                        "volume": 0,
                        "dimensoes": dims,
                        "peso": 0.0,
                        "iva": 23.0,
                        "total": 0.0
                    })
                    continue
                except ValueError as e:
                    print(f"⚠️ Erro conversão quantidade: {e}")
                    pass
        
        # Estratégia 2 (fallback): Regex original para formatos simples
        produto_match = re.match(
            r'^([A-Z0-9]{8,})\s+'
            r'(.+?)\s+'
            r'([\d,\.]+)\s+'
            r'([A-Z]{2,4})(?:\s|$)',
            stripped,
            re.IGNORECASE
        )
        
        if produto_match:
            codigo = produto_match.group(1).strip()
            descricao = produto_match.group(2).strip()
            quantidade_str = produto_match.group(3).strip()
            unidade = produto_match.group(4).strip().upper()
            
            try:
                quantidade = normalize_number(quantidade_str)
            except ValueError:
                continue
            
            dims = ""
            dim_match = re.search(r'(\d{3,4})[xX×](\d{3,4})[xX×](\d{3,4})', descricao)
            if dim_match:
                dims = f"{float(dim_match.group(1))/1000:.2f}x{float(dim_match.group(2))/1000:.2f}x{float(dim_match.group(3))/1000:.2f}"
            
            produtos.append({
                "referencia_ordem": pedido_atual or "",
                "artigo": codigo,
                "descricao": descricao,
                "lote_producao": "",
                "quantidade": quantidade,
                "unidade": unidade,
                "volume": 0,
                "dimensoes": dims,
                "peso": 0.0,
                "iva": 23.0,
                "total": 0.0
            })
    
    return produtos


def parse_ordem_compra(text: str):
    """
    Parser específico para Ordens de Compra portuguesas.
    Suporta dois formatos:
    1. Linhas separadas: Referência + Descrição / Quantidade + Unidade
    2. Linha única: REFERÊNCIA DESCRIÇÃO QUANTIDADE UNIDADE DATA
    """
    produtos = []
    lines = text.split("\n")
    
    # Encontrar referências de produtos
    referencias = []
    quantidades = []
    
    for line in lines:
        stripped = line.strip()
        if not stripped:
            continue
        
        # Formato 1: Linha completa tudo junto
        # Exemplo: 26.100145COLCHAO 1,95X1,40=27"SPA CHERRY VISCO"COLMOL1.000 UN2025-10-17
        # Padrão: CÓDIGO (números.números) + DESCRIÇÃO (texto) + QUANTIDADE (número.número) + UNIDADE + DATA
        combined_match = re.match(
            r'^(\d+\.\d+)(.+?)([\d,\.]+)\s+([A-Z]{2,4})(?:\s*(\d{4}-\d{2}-\d{2}))?',
            stripped,
            re.IGNORECASE
        )
        
        if combined_match:
            codigo = combined_match.group(1)
            descricao = combined_match.group(2).strip()
            quantidade_str = combined_match.group(3)
            unidade = combined_match.group(4).upper()
            data_entrega = combined_match.group(5) if combined_match.group(5) else ""
            
            # Validar unidade
            unidades_validas = {'UN', 'UNI', 'UNID', 'PC', 'PCS', 'KG', 'G', 'M', 'M2', 'M3', 'L', 'ML', 'CX', 'PAR', 'PAC', 'SET', 'RL', 'FD'}
            if unidade in unidades_validas:
                try:
                    quantidade = normalize_number(quantidade_str)
                    
                    # Extrair dimensões
                    dims = ""
                    dim_match = re.search(r'(\d),(\d{2})[xX×](\d),(\d{2})', descricao)
                    if dim_match:
                        dims = f"{dim_match.group(1)}.{dim_match.group(2)}x{dim_match.group(3)}.{dim_match.group(4)}"
                    
                    produtos.append({
                        "artigo": codigo,
                        "descricao": descricao,
                        "quantidade": quantidade,
                        "unidade": unidade,
                        "data_entrega": data_entrega,
                        "dimensoes": dims,
                        "referencia_ordem": "",
                        "lote_producao": "",
                        "volume": 0,
                        "peso": 0.0,
                        "iva": 23.0,
                        "total": 0.0
                    })
                    continue
                except (ValueError, IndexError):
                    pass
        
        # Formato 2a: Detectar linha de quantidade + unidade PRIMEIRO (mais específico)
        # Formato: 1.000 UN 2025-10-17 [texto opcional]
        qty_match = re.match(r'^([\d,\.]+)\s+([A-Za-z]{2,4})(?:\s+(\d{4}-\d{2}-\d{2}))?', stripped)
        if qty_match:
            quantidade_str = qty_match.group(1)
            unidade = qty_match.group(2).upper()
            data_entrega = qty_match.group(3) if qty_match.group(3) else ""
            
            unidades_validas = {'UN', 'UNI', 'UNID', 'PC', 'PCS', 'KG', 'G', 'M', 'M2', 'M3', 'L', 'ML', 'CX', 'PAR', 'PAC', 'SET', 'RL', 'FD'}
            is_valid_unit = unidade in unidades_validas or data_entrega
            
            if is_valid_unit:
                try:
                    quantidade = normalize_number(quantidade_str)
                    quantidades.append({
                        'quantidade': quantidade,
                        'unidade': unidade,
                        'data_entrega': data_entrega
                    })
                    continue
                except ValueError:
                    pass
        
        # Formato 2b: Detectar linha de referência + descrição (menos específico)
        # Formato: 26.100145 COLCHAO 1,95X1,40=27"SPA CHERRY VISCO"COLMOL
        ref_match = re.match(r'^(\d+\.\d+)\s+(.+)$', stripped)
        if ref_match:
            referencias.append({
                'codigo': ref_match.group(1),
                'descricao': ref_match.group(2).strip()
            })
            continue
    
    # Se já extraiu produtos no formato 1, retornar
    if produtos:
        return produtos
    
    # Formato 2: Combinar referências com quantidades (ordem sequencial 1:1)
    if len(referencias) != len(quantidades):
        print(f"⚠️ Contagem inconsistente: {len(referencias)} referências vs {len(quantidades)} quantidades")
        print(f"   Referências: {[r['codigo'] for r in referencias]}")
        qtys_str = ["{} {}".format(q['quantidade'], q['unidade']) for q in quantidades]
        print(f"   Quantidades: {qtys_str}")
        min_count = min(len(referencias), len(quantidades))
        print(f"   Processando apenas {min_count} produtos emparelhados")
    
    paired_count = min(len(referencias), len(quantidades))
    for i in range(paired_count):
        ref = referencias[i]
        if i < len(quantidades):
            qty_info = quantidades[i]
            
            dims = ""
            dim_match = re.search(r'(\d),(\d{2})[xX×](\d),(\d{2})', ref['descricao'])
            if dim_match:
                dims = f"{dim_match.group(1)}.{dim_match.group(2)}x{dim_match.group(3)}.{dim_match.group(4)}"
            
            produtos.append({
                "artigo": ref['codigo'],
                "descricao": ref['descricao'],
                "quantidade": qty_info['quantidade'],
                "unidade": qty_info['unidade'],
                "data_entrega": qty_info['data_entrega'],
                "dimensoes": dims,
                "referencia_ordem": "",
                "lote_producao": "",
                "volume": 0,
                "peso": 0.0,
                "iva": 23.0,
                "total": 0.0
            })
    
    return produtos


def parse_bon_commande(text: str):
    """
    Parser dedicado para BON DE COMMANDE (Notas de Encomenda francesas).
    
    Formato esperado:
    Désignation  Quantité  Prix unitaire  Montant
    MATELAS SAN REMO 140x190  2 202.00€ 404.00€
    
    Extrai:
    - Referência/Designação do produto
    - Quantidade
    - Preço unitário
    - Total da linha
    """
    produtos = []
    lines = text.split("\n")
    
    # Buscar cliente
    cliente = ""
    cliente_match = re.search(r'ADRESSE DE LIVRAISON\s+([^\n]+)', text, re.IGNORECASE)
    if cliente_match:
        cliente = cliente_match.group(1).strip()
    
    # Buscar data
    data = ""
    data_match = re.search(r'DATE\s*:\s*(\d{2}\.\d{2}\.\d{2})', text, re.IGNORECASE)
    if data_match:
        data = data_match.group(1)
    
    # Buscar contremarque
    contremarque = ""
    cm_match = re.search(r'CONTREMARQUE\s*:\s*([^\n]+)', text, re.IGNORECASE)
    if cm_match:
        contremarque = cm_match.group(1).strip()
    
    in_product_section = False
    product_header_found = False
    
    # Buffer para juntar linhas (PyMuPDF coloca cada campo numa linha separada)
    i = 0
    while i < len(lines):
        stripped = lines[i].strip()
        if not stripped:
            i += 1
            continue
        
        # Detectar início da seção de produtos (palavras-chave podem estar em linhas separadas)
        if re.search(r'Désignation|Quantité|Prix\s+unitaire', stripped, re.IGNORECASE):
            product_header_found = True
            i += 1
            continue
        
        # Skip header lines
        if re.search(r'^(Montant|STOCK\s+MAGASIN)$', stripped, re.IGNORECASE):
            in_product_section = True
            i += 1
            continue
        
        # Detectar fim da seção
        if re.search(r'^TOTAL|^ADRESSE|^BON DE COMMANDE|^Livraison|^AIRE DES|^Tél|^SIREN', stripped, re.IGNORECASE):
            in_product_section = False
            i += 1
            continue
        
        if in_product_section and product_header_found:
            # Formato OCR: cada campo numa linha separada
            # Linha 1: MATELAS SAN REMO 140x190
            # Linha 2: 2
            # Linha 3: 202.00€
            # Linha 4: 404.00€
            
            # Verificar se esta é uma linha de produto (descrição)
            if i + 3 < len(lines):
                desc_line = lines[i].strip()
                qty_line = lines[i+1].strip()
                price_line = lines[i+2].strip()
                total_line = lines[i+3].strip()
                
                # Verificar se próximas 3 linhas parecem quantidade + preços
                if (re.match(r'^\d+$', qty_line) and
                    re.match(r'^[\d,\.]+\s*€', price_line) and
                    re.match(r'^[\d,\.]+\s*€', total_line)):
                    
                    try:
                        quantidade = int(qty_line)
                        preco_str = price_line.replace('€', '').strip()
                        total_str = total_line.replace('€', '').strip()
                        
                        preco_unitario = normalize_number(preco_str)
                        total_produto = normalize_number(total_str)
                        
                        # Extrair dimensões
                        dims = ""
                        dim_match = re.search(r'(\d{2,3})\s*[xX×]\s*(\d{2,3})', desc_line)
                        if dim_match:
                            dims = f"{dim_match.group(1)}x{dim_match.group(2)}"
                        
                        # Extrair código (ex: SAN REMO, RIVIERA)
                        codigo = ""
                        cod_match = re.match(r'^MATELAS\s+([A-Z\s]+?)\s+\d', desc_line, re.IGNORECASE)
                        if cod_match:
                            codigo = cod_match.group(1).strip()
                        
                        produtos.append({
                            "artigo": codigo if codigo else desc_line[:20],
                            "descricao": desc_line,
                            "quantidade": float(quantidade),
                            "unidade": "UN",
                            "preco_unitario": preco_unitario,
                            "total": total_produto,
                            "dimensoes": dims,
                            "cliente": cliente,
                            "data_encomenda": data,
                            "contremarque": contremarque,
                            "referencia_ordem": "",
                            "lote_producao": "",
                            "volume": 0,
                            "peso": 0.0,
                            "iva": 20.0
                        })
                        
                        print(f"✅ Produto Bon de Commande: {desc_line} - {quantidade} UN")
                        i += 4  # Pular as 4 linhas processadas
                        continue
                    except (ValueError, IndexError) as e:
                        pass
        
        i += 1
    
    return produtos


def parse_pedido_espanhol(text: str):
    """
    Parser dedicado para PEDIDO espanhol (NATURCOLCHON, COSGUI, etc).
    
    Formatos esperados:
    1. NATURCOLCHON: Código | Descripción | Unidades | Precio | Importe
       COPR1520 COLCHON PRAGA DE 150X200 CM 5,00 175,00 875,00
    
    2. COSGUI: Código | Descripción | Cantidad
       LUSTOPVS135190 COLCHON TOP VISCO 2019 135X190 4,00
    """
    produtos = []
    lines = text.split("\n")
    
    # Buscar número de pedido
    pedido_num = ""
    ped_match = re.search(r'(?:Pedido|Número).*?(\d+)', text, re.IGNORECASE)
    if ped_match:
        pedido_num = ped_match.group(1)
    
    # Buscar data
    fecha = ""
    fecha_match = re.search(r'Fecha.*?(\d{2}/\d{2}/\d{4})', text, re.IGNORECASE)
    if fecha_match:
        fecha = fecha_match.group(1)
    
    # Buscar proveedor
    proveedor = ""
    prov_match = re.search(r'Proveedor.*?([A-Z\s\.]+)', text, re.IGNORECASE)
    if prov_match:
        proveedor = prov_match.group(1).strip()
    
    in_product_section = False
    
    # Multi-line buffer: tentar juntar 3 linhas para formato COSGUI (qty, desc, code em linhas separadas)
    i = 0
    while i < len(lines):
        stripped = lines[i].strip()
        if not stripped:
            i += 1
            continue
        
        # Detectar início da seção de produtos (keywords podem vir em linhas separadas)
        if re.search(r'Artículo|Descripción|Cantidad|Código', stripped, re.IGNORECASE):
            in_product_section = True
            i += 1
            continue
        
        # Detectar fim da seção
        if re.search(r'^Total|^Importe neto|^Notas|^Plazo|^Base I\.V\.A', stripped, re.IGNORECASE):
            in_product_section = False
            i += 1
            continue
        
        # NOVO: Buffer multi-linha para formato COSGUI (quantidade, descrição, código em linhas separadas)
        # Tentar juntar próximas 3 linhas se parecerem ser: QTY + DESC + CODE
        if i + 2 < len(lines):
            line1 = lines[i].strip()
            line2 = lines[i+1].strip()
            line3 = lines[i+2].strip()
            
            # Padrão: linha1=quantidade, linha2=descrição, linha3=código
            if (re.match(r'^[\d,]+$', line1) and  # Quantidade pura
                len(line2) > 10 and  # Descrição tem texto
                re.match(r'^[A-Z0-9]{6,}$', line3)):  # Código alfanumérico
                
                # VALIDAÇÕES ANTI-FALSO-POSITIVO:
                # 1. Código não pode ser número puro (evita números de documento)
                if re.match(r'^\d+$', line3):
                    i += 1
                    continue
                
                # 2. Código não pode começar com PT (evita NIFs portugueses)
                if line3.startswith('PT'):
                    i += 1
                    continue
                
                # 3. Quantidade não pode ser muito alta (evita telefones/códigos postais)
                try:
                    qty_check = normalize_number(line1)
                    if qty_check > 100:  # Produtos geralmente < 100 unidades
                        i += 1
                        continue
                except:
                    pass
                
                # 4. Descrição não pode conter palavras de endereço
                address_words = ['POLIGONO', 'NAVE', 'CALLE', 'RUA', 'AVENIDA', 'ZONA', 'INDUSTRIAL', 'MORERO', 'GUARNIZO']
                desc_upper = line2.upper()
                if any(word in desc_upper for word in address_words):
                    i += 1
                    continue
                
                # 5. Linha 3 (código) não pode ter palavras (evita "GUARNIZO", "PORTUGAL", "POLIGONO", etc como código)
                if any(word in line3.upper() for word in ['GUARNIZO', 'PORTUGAL', 'ESPAÑA', 'FRANCE', 'ADMINISTRA', 'POLIGONO', 'INDUTRIAL', 'MORERO']):
                    i += 1
                    continue
                
                # Reconstruir linha no formato esperado: CÓDIGO DESCRIPCIÓN CANTIDAD
                reconstructed = f"{line3} {line2} {line1}"
                print(f"🔧 Buffer multi-linha: '{line1}' + '{line2}' + '{line3}' → '{reconstructed}'")
                
                # Tentar match no formato 2
                match2 = re.match(
                    r'^([A-Z0-9]{6,})\s+(.+?)\s+([\d,]+)$',
                    reconstructed
                )
                
                if match2:
                    codigo = match2.group(1)
                    descripcion = match2.group(2).strip()
                    cantidad_str = match2.group(3)
                    
                    try:
                        cantidad = normalize_number(cantidad_str)
                        
                        # Extrair dimensões
                        dims = ""
                        dim_match = re.search(r'(\d{2,3})[xX×](\d{2,3})', descripcion)
                        if dim_match:
                            dims = f"{dim_match.group(1)}x{dim_match.group(2)}"
                        
                        produtos.append({
                            "artigo": codigo,
                            "descricao": descripcion,
                            "quantidade": cantidad,
                            "unidade": "UN",
                            "preco_unitario": 0.0,
                            "total": 0.0,
                            "dimensoes": dims,
                            "pedido_numero": pedido_num,
                            "fecha": fecha,
                            "proveedor": proveedor,
                            "referencia_ordem": "",
                            "lote_producao": "",
                            "volume": 0,
                            "peso": 0.0,
                            "iva": 21.0
                        })
                        print(f"✅ Produto multi-linha extraído: {codigo} - {descripcion} - {cantidad}")
                        i += 3  # Pular as 3 linhas processadas
                        continue
                    except ValueError:
                        pass
        
        if in_product_section or True:  # SEMPRE tentar parsear (headers podem vir depois)
            # Formato 1B: DESCRIPCIÓN CÓDIGO TOTAL PRECIO UNIDADES (formato invertido NATURCOLCHON)
            # Exemplo: COLCHON PRAGA DE 150X200 CM*NUEVO* COPR1520 875,00 175,00 5,00
            # VERIFICAR PRIMEIRO pois tem 3 números (mais específico)
            match1b = re.match(
                r'^(.+?)\s+([A-Z0-9]{4,})\s+([\d,]+)\s+([\d,]+)\s+([\d,]+)$',
                stripped
            )
            
            # Formato 1: CÓDIGO DESCRIPCIÓN UNIDADES PRECIO IMPORTE
            # Exemplo: COPR1520 COLCHON PRAGA DE 150X200 CM*NUEVO* 5,00 175,00 875,00
            match1 = re.match(
                r'^([A-Z0-9]{4,})\s+(.+?)\s+([\d,]+)\s+([\d,]+)\s+([\d,]+)$',
                stripped
            )
            
            if match1b:
                # Formato invertido: descrição vem primeiro
                descripcion = match1b.group(1).strip()
                codigo = match1b.group(2)
                total_str = match1b.group(3).replace(',', '.')
                precio_str = match1b.group(4).replace(',', '.')
                cantidad_str = match1b.group(5).replace(',', '.')
                
                # VALIDAÇÕES ANTI-FALSO-POSITIVO (igual buffer multi-linha)
                is_valid = True
                # 1. Código não pode ser número puro
                if re.match(r'^\d+$', codigo):
                    is_valid = False
                # 2. Código não pode começar com PT (NIFs)
                if codigo.startswith('PT'):
                    is_valid = False
                # 3. Quantidade não pode ser > 100
                try:
                    if float(cantidad_str) > 100:
                        is_valid = False
                except:
                    pass
                # 4. Descrição não pode ter palavras de endereço
                address_words = ['POLIGONO', 'NAVE', 'CALLE', 'RUA', 'AVENIDA', 'ZONA', 'INDUSTRIAL', 'MORERO', 'GUARNIZO']
                if any(word in descripcion.upper() for word in address_words):
                    is_valid = False
                # 5. Código não pode ter palavras de endereço
                if any(word in codigo.upper() for word in ['POLIGONO', 'INDUTRIAL', 'MORERO', 'GUARNIZO', 'PORTUGAL', 'ESPAÑA', 'ADMINISTRA']):
                    is_valid = False
                
                if is_valid:
                    try:
                        cantidad = float(cantidad_str)
                        precio = float(precio_str)
                        total = float(total_str)
                        
                        # Extrair dimensões
                        dims = ""
                        dim_match = re.search(r'(\d{2,3})[xX×](\d{2,3})', descripcion)
                        if dim_match:
                            dims = f"{dim_match.group(1)}x{dim_match.group(2)}"
                        
                        produtos.append({
                            "artigo": codigo,
                            "descricao": descripcion,
                            "quantidade": cantidad,
                            "unidade": "UN",
                            "preco_unitario": precio,
                            "total": total,
                            "dimensoes": dims,
                            "pedido_numero": pedido_num,
                            "fecha": fecha,
                            "proveedor": proveedor,
                            "referencia_ordem": "",
                            "lote_producao": "",
                            "volume": 0,
                            "peso": 0.0,
                            "iva": 21.0  # IVA Espanha padrão
                        })
                        print(f"✅ Formato 1B extraído: {codigo} - {descripcion} - {cantidad}")
                        i += 1
                        continue
                    except ValueError:
                        pass
                # Se inválido ou falhou, pular linha
                i += 1
                continue
            
            elif match1:
                codigo = match1.group(1)
                descripcion = match1.group(2).strip()
                cantidad_str = match1.group(3).replace(',', '.')
                precio_str = match1.group(4).replace(',', '.')
                total_str = match1.group(5).replace(',', '.')
                
                # VALIDAÇÕES ANTI-FALSO-POSITIVO (igual buffer multi-linha)
                # 1. Código não pode ser número puro
                if re.match(r'^\d+$', codigo):
                    i += 1
                    continue
                # 2. Código não pode começar com PT (NIFs)
                if codigo.startswith('PT'):
                    i += 1
                    continue
                # 3. Quantidade não pode ser > 100
                try:
                    if float(cantidad_str) > 100:
                        i += 1
                        continue
                except:
                    pass
                # 4. Descrição não pode ter palavras de endereço
                address_words = ['POLIGONO', 'NAVE', 'CALLE', 'RUA', 'AVENIDA', 'ZONA', 'INDUSTRIAL', 'MORERO', 'GUARNIZO']
                if any(word in descripcion.upper() for word in address_words):
                    i += 1
                    continue
                # 5. Código não pode ter palavras de endereço
                if any(word in codigo.upper() for word in ['POLIGONO', 'INDUTRIAL', 'MORERO', 'GUARNIZO', 'PORTUGAL', 'ESPAÑA', 'ADMINISTRA']):
                    i += 1
                    continue
                
                try:
                    cantidad = float(cantidad_str)
                    precio = float(precio_str)
                    total = float(total_str)
                    
                    # Extrair dimensões
                    dims = ""
                    dim_match = re.search(r'(\d{2,3})[xX×](\d{2,3})', descripcion)
                    if dim_match:
                        dims = f"{dim_match.group(1)}x{dim_match.group(2)}"
                    
                    produtos.append({
                        "artigo": codigo,
                        "descricao": descripcion,
                        "quantidade": cantidad,
                        "unidade": "UN",
                        "preco_unitario": precio,
                        "total": total,
                        "dimensoes": dims,
                        "pedido_numero": pedido_num,
                        "fecha": fecha,
                        "proveedor": proveedor,
                        "referencia_ordem": "",
                        "lote_producao": "",
                        "volume": 0,
                        "peso": 0.0,
                        "iva": 21.0  # IVA Espanha padrão
                    })
                    print(f"✅ Formato 1 extraído: {codigo} - {descripcion} - {cantidad}")
                    continue
                except ValueError:
                    pass
            
            # Formato 2: CÓDIGO DESCRIPCIÓN CANTIDAD
            # Exemplo: LUSTOPVS135190 COLCHON TOP VISCO 2019 135X190 4,00
            match2 = re.match(
                r'^([A-Z0-9]{6,})\s+(.+?)\s+([\d,]+)$',
                stripped
            )
            
            if match2:
                codigo = match2.group(1)
                descripcion = match2.group(2).strip()
                cantidad_str = match2.group(3).replace(',', '.')
                
                # VALIDAÇÕES (mesmas dos outros formatos)
                is_valid = True
                if re.match(r'^\d+$', codigo):
                    is_valid = False
                if codigo.startswith('PT'):
                    is_valid = False
                try:
                    if float(cantidad_str) > 100:
                        is_valid = False
                except:
                    pass
                address_words = ['POLIGONO', 'NAVE', 'CALLE', 'RUA', 'AVENIDA', 'ZONA', 'INDUSTRIAL', 'MORERO', 'GUARNIZO']
                if any(word in descripcion.upper() for word in address_words):
                    is_valid = False
                if any(word in codigo.upper() for word in ['POLIGONO', 'INDUTRIAL', 'MORERO', 'GUARNIZO', 'PORTUGAL', 'ESPAÑA', 'ADMINISTRA']):
                    is_valid = False
                
                if not is_valid:
                    i += 1
                    continue
                
                try:
                    cantidad = float(cantidad_str)
                    
                    # Extrair dimensões
                    dims = ""
                    dim_match = re.search(r'(\d{2,3})[xX×](\d{2,3})', descripcion)
                    if dim_match:
                        dims = f"{dim_match.group(1)}x{dim_match.group(2)}"
                    
                    produtos.append({
                        "artigo": codigo,
                        "descricao": descripcion,
                        "quantidade": cantidad,
                        "unidade": "UN",
                        "preco_unitario": 0.0,
                        "total": 0.0,
                        "dimensoes": dims,
                        "pedido_numero": pedido_num,
                        "fecha": fecha,
                        "proveedor": proveedor,
                        "referencia_ordem": "",
                        "lote_producao": "",
                        "volume": 0,
                        "peso": 0.0,
                        "iva": 21.0
                    })
                except ValueError:
                    pass
        
        # Incrementar contador (se não houve continue antes)
        i += 1
    
    return produtos


# ============================================================================
# FUNÇÕES UNIVERSAIS DE EXTRAÇÃO (Fuzzy Matching + Table Extraction)
# ============================================================================

def universal_kv_extract(text: str, file_path: str = None):
    """
    Extração universal de key-value pairs usando fuzzy matching (rapidfuzz).
    Encontra: fornecedor/supplier/proveedor, NIF, IBAN, número documento, data, número encomenda
    """
    if not RAPIDFUZZ_AVAILABLE:
        return {}
    
    result = {}
    lines = text.split('\n')
    
    # Sinônimos multi-idioma para campos-chave
    synonyms = {
        'fornecedor': ['fornecedor', 'supplier', 'proveedor', 'empresa', 'vendedor', 'seller', 'company'],
        'nif': ['nif', 'vat', 'nipc', 'tax id', 'cif', 'dni', 'identificação fiscal'],
        'iban': ['iban', 'account', 'conta bancária', 'bank account', 'cuenta bancaria'],
        'documento': ['documento', 'document', 'factura', 'fatura', 'invoice', 'guia', 'pedido', 'order'],
        'data': ['data', 'date', 'fecha', 'datum'],
        'encomenda': ['encomenda', 'order', 'pedido', 'purchase order', 'po', 'commande']
    }
    
    for line in lines:
        line_clean = line.strip()
        if not line_clean or len(line_clean) < 3:
            continue
        
        # Split em possível key:value
        parts = line_clean.split(':', 1)
        if len(parts) == 2:
            key_candidate = parts[0].strip().lower()
            value_candidate = parts[1].strip()
            
            # Fuzzy match para cada categoria
            for field, variants in synonyms.items():
                if field in result:  # Já encontrado
                    continue
                
                # Usa rapidfuzz para encontrar melhor match
                best_match = process.extractOne(key_candidate, variants, scorer=fuzz.ratio)
                
                if best_match and best_match[1] >= 70:  # Score >= 70%
                    result[field] = value_candidate
                    break
    
    return result


def universal_table_extract(file_path: str):
    """
    Extração universal de tabelas usando Camelot + pdfplumber.
    Retorna lista de produtos extraídos de tabelas detectadas.
    """
    produtos = []
    
    # Método 1: Camelot (melhor para tabelas com bordas)
    if CAMELOT_AVAILABLE and file_path.lower().endswith('.pdf'):
        try:
            tables = camelot.read_pdf(file_path, pages='all', flavor='lattice')
            
            if len(tables) > 0:
                print(f"✅ Camelot detectou {len(tables)} tabela(s)")
                
                for table_idx, table in enumerate(tables):
                    df = table.df
                    
                    # Tenta identificar colunas de produto (heurística)
                    possible_headers = df.iloc[0].tolist() if len(df) > 0 else []
                    header_lower = [str(h).lower() for h in possible_headers]
                    
                    # Procura colunas importantes
                    col_map = {}
                    for idx, h in enumerate(header_lower):
                        if any(kw in h for kw in ['código', 'codigo', 'code', 'ref', 'artigo', 'article']):
                            col_map['codigo'] = idx
                        elif any(kw in h for kw in ['descrição', 'descripcion', 'description', 'designation', 'produto']):
                            col_map['descricao'] = idx
                        elif any(kw in h for kw in ['quantidade', 'qty', 'qtd', 'quant', 'unidades', 'cantidad']):
                            col_map['quantidade'] = idx
                        elif any(kw in h for kw in ['preço', 'precio', 'price', 'unitário', 'unit']):
                            col_map['preco'] = idx
                    
                    # Extrai produtos
                    for row_idx in range(1, len(df)):
                        row = df.iloc[row_idx]
                        
                        produto = {}
                        if 'codigo' in col_map:
                            produto['artigo'] = str(row[col_map['codigo']]).strip()
                        if 'descricao' in col_map:
                            produto['descricao'] = str(row[col_map['descricao']]).strip()
                        if 'quantidade' in col_map:
                            qty_str = str(row[col_map['quantidade']]).strip()
                            try:
                                produto['quantidade'] = float(qty_str.replace(',', '.'))
                            except:
                                produto['quantidade'] = 0.0
                        if 'preco' in col_map:
                            preco_str = str(row[col_map['preco']]).strip()
                            try:
                                produto['preco_unitario'] = float(preco_str.replace(',', '.'))
                            except:
                                produto['preco_unitario'] = 0.0
                        
                        # Valida produto mínimo (tem código OU descrição + quantidade)
                        if (produto.get('artigo') or produto.get('descricao')) and produto.get('quantidade', 0) > 0:
                            produtos.append(produto)
        
        except Exception as e:
            print(f"⚠️ Camelot falhou: {e}")
    
    # Método 2: pdfplumber (melhor para tabelas sem bordas)
    if PDFPLUMBER_AVAILABLE and file_path.lower().endswith('.pdf') and len(produtos) == 0:
        try:
            with pdfplumber.open(file_path) as pdf:
                for page in pdf.pages:
                    tables = page.extract_tables()
                    
                    if tables:
                        print(f"✅ pdfplumber detectou {len(tables)} tabela(s) na página {page.page_number}")
                        
                        for table in tables:
                            if not table or len(table) < 2:
                                continue
                            
                            # Primeira linha = headers
                            headers = [str(h).lower().strip() if h else '' for h in table[0]]
                            
                            # Mapeia colunas
                            col_map = {}
                            for idx, h in enumerate(headers):
                                if any(kw in h for kw in ['código', 'codigo', 'code', 'ref', 'artigo']):
                                    col_map['codigo'] = idx
                                elif any(kw in h for kw in ['descrição', 'descripcion', 'description', 'designation']):
                                    col_map['descricao'] = idx
                                elif any(kw in h for kw in ['quantidade', 'qty', 'qtd', 'quant', 'unidades', 'cantidad']):
                                    col_map['quantidade'] = idx
                                elif any(kw in h for kw in ['preço', 'precio', 'price', 'unitário', 'unit']):
                                    col_map['preco'] = idx
                            
                            # Extrai linhas
                            for row in table[1:]:
                                if not row or len(row) == 0:
                                    continue
                                
                                produto = {}
                                if 'codigo' in col_map and col_map['codigo'] < len(row):
                                    produto['artigo'] = str(row[col_map['codigo']]).strip() if row[col_map['codigo']] else ''
                                if 'descricao' in col_map and col_map['descricao'] < len(row):
                                    produto['descricao'] = str(row[col_map['descricao']]).strip() if row[col_map['descricao']] else ''
                                if 'quantidade' in col_map and col_map['quantidade'] < len(row):
                                    qty_str = str(row[col_map['quantidade']]).strip() if row[col_map['quantidade']] else '0'
                                    try:
                                        produto['quantidade'] = float(qty_str.replace(',', '.'))
                                    except:
                                        produto['quantidade'] = 0.0
                                if 'preco' in col_map and col_map['preco'] < len(row):
                                    preco_str = str(row[col_map['preco']]).strip() if row[col_map['preco']] else '0'
                                    try:
                                        produto['preco_unitario'] = float(preco_str.replace(',', '.'))
                                    except:
                                        produto['preco_unitario'] = 0.0
                                
                                # Valida produto mínimo
                                if (produto.get('artigo') or produto.get('descricao')) and produto.get('quantidade', 0) > 0:
                                    produtos.append(produto)
        
        except Exception as e:
            print(f"⚠️ pdfplumber falhou: {e}")
    
    # Filtrar produtos inválidos (endereços, códigos postais, etc)
    produtos_validos = []
    palavras_invalidas = {
        'tertres', 'moissons', 'maxiliterie', 'colmol', 'adresse', 'livraison',
        'zone', 'commercial', 'rue', 'rua', 'avenida', 'street', 'avenue',
        'cidade', 'city', 'codigo', 'postal', 'telefone', 'phone', 'tel',
        'fax', 'email', 'e-mail', 'cliente', 'customer', 'fornecedor', 'supplier'
    }
    
    for p in produtos:
        artigo = str(p.get('artigo', '')).lower()
        descricao = str(p.get('descricao', '')).lower()
        quantidade = p.get('quantidade', 0)
        
        # Filtro 1: Evitar códigos postais como quantidades (>= 5 dígitos)
        if quantidade > 9999:
            continue
        
        # Filtro 2: Evitar palavras de endereço no artigo ou descrição
        texto_completo = f"{artigo} {descricao}"
        if any(palavra in texto_completo for palavra in palavras_invalidas):
            continue
        
        # Filtro 3: Artigo muito curto ou genérico
        if artigo and len(artigo) < 2:
            continue
        
        # Produto válido
        produtos_validos.append(p)
    
    if produtos_validos:
        print(f"✅ Extração universal de tabelas: {len(produtos_validos)} produtos válidos (filtrados {len(produtos) - len(produtos_validos)})")
    
    return produtos_validos


def parse_generic_document(text: str, file_path: str = None):
    """
    Parser genérico universal - última tentativa quando parsers específicos falharem.
    Combina regex heurísticos + table extraction + fuzzy matching.
    """
    produtos = []
    metadata = {}
    
    # 1. Extração de metadados com fuzzy matching
    if file_path:
        metadata = universal_kv_extract(text, file_path)
        print(f"📋 Metadados extraídos (fuzzy): {list(metadata.keys())}")
    
    # 2. Tentativa de extração por tabelas
    if file_path:
        produtos = universal_table_extract(file_path)
    
    # 3. Se ainda não tem produtos, tenta regex genéricos
    if len(produtos) == 0:
        lines = text.split('\n')
        
        # Regex genérico para linhas de produto (artigo + descrição + quantidade + preço)
        generic_patterns = [
            # Padrão 1: CÓDIGO DESCRIÇÃO QTY PREÇO
            r'^\s*([A-Z0-9\-]+)\s+(.{10,60}?)\s+(\d+[,.]?\d*)\s+(\d+[,.]?\d+)\s*$',
            # Padrão 2: CÓDIGO | DESCRIÇÃO | QTY
            r'^\s*([A-Z0-9\-]+)\s*\|\s*(.{10,60}?)\s*\|\s*(\d+[,.]?\d*)',
            # Padrão 3: QTY DESCRIÇÃO CÓDIGO
            r'^\s*(\d+[,.]?\d*)\s+(.{10,60}?)\s+([A-Z0-9\-]+)\s*$'
        ]
        
        for line in lines:
            line_stripped = line.strip()
            if len(line_stripped) < 10:
                continue
            
            for pattern_idx, pattern in enumerate(generic_patterns):
                match = re.match(pattern, line_stripped)
                if match:
                    try:
                        if pattern_idx == 0:  # CÓDIGO DESC QTY PREÇO
                            codigo, desc, qty, preco = match.groups()
                            produtos.append({
                                'artigo': codigo.strip(),
                                'descricao': desc.strip(),
                                'quantidade': float(qty.replace(',', '.')),
                                'preco_unitario': float(preco.replace(',', '.'))
                            })
                        elif pattern_idx == 1:  # CÓDIGO | DESC | QTY
                            codigo, desc, qty = match.groups()
                            produtos.append({
                                'artigo': codigo.strip(),
                                'descricao': desc.strip(),
                                'quantidade': float(qty.replace(',', '.')),
                                'preco_unitario': 0.0
                            })
                        elif pattern_idx == 2:  # QTY DESC CÓDIGO
                            qty, desc, codigo = match.groups()
                            produtos.append({
                                'artigo': codigo.strip(),
                                'descricao': desc.strip(),
                                'quantidade': float(qty.replace(',', '.')),
                                'preco_unitario': 0.0
                            })
                        break  # Encontrou match, próxima linha
                    except ValueError:
                        continue
    
    if produtos:
        print(f"✅ Parser genérico universal: {len(produtos)} produtos extraídos")
    else:
        print("⚠️ Parser genérico universal: 0 produtos extraídos")
    
    return {'produtos': produtos, 'metadata': metadata}


def parse_portuguese_document(text: str, qr_codes=None, texto_pdfplumber_curto=False, file_path=None):
    """
    Extrai cabeçalho (req/doc/fornecedor/data) e linhas de produto.
    Usa fallback universal (fuzzy matching + table extraction) se parsers específicos falharem.
    """
    if qr_codes is None:
        qr_codes = []
    
    doc_type = detect_document_type(text)
    print(f"📄 Tipo de documento detectado: {doc_type}")

    lines = text.split("\n")
    result = {
        "numero_requisicao": "",
        "document_number": "",
        "po_number": "",
        "supplier_name": "",
        "delivery_date": "",
        "qr_codes": qr_codes,
        "produtos": [],
        "lines": [],
        "totals": {
            "total_lines": 0,
            "total_quantity": 0
        },
        "tipo_documento": doc_type,
        "texto_completo": text,
        "baixa_qualidade_texto": texto_pdfplumber_curto,
    }

    patterns = {
        "req": r"(?:req|requisição)\.?\s*n?[oº]?\s*:?\s*([A-Z0-9\-/]+)",
        "doc": r"(?:guia|gr|documento|fatura)\.?\s*n?[oº]?\s*:?\s*([A-Z0-9\-/]+)",
        "data": r"(\d{1,2}[/-]\d{1,2}[/-]\d{2,4})",
        "fornecedor": r"(?:fornecedor|empresa)\.?\s*:?\s*([^\n]+)",
    }

    for ln in lines:
        low = ln.lower().strip()

        if not result["numero_requisicao"]:
            m = re.search(patterns["req"], low, re.IGNORECASE)
            if m:
                result["numero_requisicao"] = m.group(1).upper()

        if not result["document_number"]:
            m = re.search(patterns["doc"], low, re.IGNORECASE)
            if m:
                result["document_number"] = m.group(1).upper()

        if not result["delivery_date"]:
            m = re.search(patterns["data"], ln)
            if m:
                result["delivery_date"] = m.group(1)

        if not result["supplier_name"]:
            m = re.search(patterns["fornecedor"], low)
            if m:
                result["supplier_name"] = m.group(1).title()

    if doc_type == "PEDIDO_ESPANHOL":
        produtos = parse_pedido_espanhol(text)
        if produtos:
            result["produtos"] = produtos
            print(f"✅ Extraídos {len(produtos)} produtos do Pedido Espanhol")
            
            # Extrair metadados dos produtos
            if produtos and produtos[0].get("proveedor"):
                result["supplier_name"] = produtos[0]["proveedor"]
            if produtos and produtos[0].get("fecha"):
                result["delivery_date"] = produtos[0]["fecha"]
            if produtos and produtos[0].get("pedido_numero"):
                result["document_number"] = produtos[0]["pedido_numero"]
        else:
            print("⚠️ Parser Pedido Espanhol retornou 0 produtos")
    elif doc_type == "BON_COMMANDE":
        produtos = parse_bon_commande(text)
        if produtos:
            result["produtos"] = produtos
            print(f"✅ Extraídos {len(produtos)} produtos do Bon de Commande")
            
            # Extrair cliente e data dos produtos
            if produtos and produtos[0].get("cliente"):
                result["supplier_name"] = produtos[0]["cliente"]
            if produtos and produtos[0].get("data_encomenda"):
                result["delivery_date"] = produtos[0]["data_encomenda"]
            
            # Extrair contremarque como número de documento
            if produtos and produtos[0].get("contremarque"):
                result["document_number"] = produtos[0]["contremarque"]
        else:
            print("⚠️ Parser Bon de Commande retornou 0 produtos")
    elif doc_type == "ORDEM_COMPRA":
        produtos = parse_ordem_compra(text)
        if produtos:
            result["produtos"] = produtos
            print(f"✅ Extraídos {len(produtos)} produtos da Ordem de Compra")
            
            # Extrair número da ordem de compra
            oc_match = re.search(r'ORDEM\s+COMPRA\s+N[ºo]?\s*([A-Z0-9]+)', text, re.IGNORECASE)
            if oc_match:
                result["po_number"] = oc_match.group(1)
                result["document_number"] = oc_match.group(1)
        else:
            print("⚠️ Parser Ordem de Compra retornou 0 produtos")
    elif doc_type == "FATURA_ELASTRON":
        produtos = parse_fatura_elastron(text)
        if produtos:
            result["produtos"] = produtos
            result["supplier_name"] = "Elastron Portugal, SA"
            print(f"✅ Extraídos {len(produtos)} produtos da Fatura Elastron")
        else:
            print("⚠️ Parser Elastron retornou 0 produtos, tentando parser genérico...")
            produtos = parse_guia_generica(text)
            if produtos:
                result["produtos"] = produtos
                print(f"✅ Extraídos {len(produtos)} produtos com parser genérico")
    elif doc_type == "GUIA_COLMOL":
        produtos = parse_guia_colmol(text)
        if produtos:
            result["produtos"] = produtos
            result["supplier_name"] = "Colmol - Colchões S.A"
            print(f"✅ Extraídos {len(produtos)} produtos da Guia Colmol")
        else:
            print("⚠️ Parser Colmol retornou 0 produtos, tentando parser genérico...")
            produtos = parse_guia_generica(text)
            if produtos:
                result["produtos"] = produtos
                print(f"✅ Extraídos {len(produtos)} produtos com parser genérico")
    else:
        if "GUIA" in doc_type:
            produtos = parse_guia_generica(text)
            if produtos:
                result["produtos"] = produtos
                print(f"✅ Extraídos {len(produtos)} produtos com parser genérico de guias")
        
        if not result.get("produtos"):
            guia_products = extract_guia_remessa_products(text)
            if guia_products:
                result["produtos"] = guia_products
            else:
                product_lines = extract_product_lines(text)
                legacy = []
                for p in product_lines:
                    legacy.append({
                        "supplier_code": p["codigo_fornecedor"],
                        "description": p["descricao"],
                        "linha_raw": p["linha_raw"],
                        "unit": p["unidade"],
                        "qty": p["quantidade"],
                        "mini_codigo": p["mini_codigo"],
                        "dimensoes": p["dimensoes"],
                    })
                result["lines"] = legacy

    # FALLBACK UNIVERSAL: Se nenhum parser específico extraiu produtos, usa extração universal
    if not result["produtos"] and file_path:
        print("🔄 Nenhum parser específico funcionou - tentando extração universal...")
        generic_result = parse_generic_document(text, file_path)
        
        if generic_result.get('produtos'):
            result["produtos"] = generic_result['produtos']
            print(f"✅ Extração universal bem-sucedida: {len(result['produtos'])} produtos")
            
            # Atualiza metadados se disponível
            if generic_result.get('metadata'):
                metadata = generic_result['metadata']
                if not result["supplier_name"] and metadata.get('fornecedor'):
                    result["supplier_name"] = metadata['fornecedor']
                if not result["document_number"] and metadata.get('documento'):
                    result["document_number"] = metadata['documento']
                if not result["delivery_date"] and metadata.get('data'):
                    result["delivery_date"] = metadata['data']
                if not result["po_number"] and metadata.get('encomenda'):
                    result["po_number"] = metadata['encomenda']

    if result["produtos"]:
        result["totals"]["total_lines"] = len(result["produtos"])
        result["totals"]["total_quantity"] = sum(p.get("quantidade", 0) for p in result["produtos"])
        
        if not result["po_number"] and result["produtos"]:
            for produto in result["produtos"]:
                ref = produto.get("referencia_ordem", "")
                po_match = re.match(r'^([A-Z0-9]+)\s+[NnºN]', ref, re.IGNORECASE)
                if po_match:
                    result["po_number"] = po_match.group(1).upper()
                    break
    elif result["lines"]:
        result["totals"]["total_lines"] = len(result["lines"])
        result["totals"]["total_quantity"] = sum(x["qty"] for x in result["lines"])

    if not result["po_number"] and result["document_number"]:
        result["po_number"] = result["document_number"]

    return result


def extract_product_lines(text: str):
    """Extrai linhas de produto com regex tolerante a formatos reais."""
    products = []
    lines = text.split("\n")

    code_pat = r"(?P<code>(?:[A-Z]{1}[A-Z0-9\-\/\.]{2,}))"  # BLC-D25-200x300, REF-123, etc.
    dens_pat = r"(?P<densidade>D\d{2})?"  # D23, D30, etc. (opcional)
    sep = r"[xX×\- ]"  # separadores
    dim_pat = rf"(?P<dim>(\d{{2,4}}){sep}(\d{{2,4}})(?:{sep}(\d{{2,4}}))?)"
    qty_pat = r"(?P<qty>\d+(?:[.,]\d+)?)\s*(?:un|uni|unid|unidades)?$"

    line_regex = re.compile(
        rf"(?i)^(?=.*{code_pat})(?=.*{dim_pat}).*{qty_pat}")

    for raw in lines:
        line = raw.strip()
        if len(line) < 5:
            continue

        m = line_regex.search(line)
        if not m:
            # Fallback: ordem trocada; procurar blocos na linha
            code_m = re.search(code_pat, line)
            dim_m = re.search(dim_pat, line)
            qty_m = re.search(qty_pat, line, flags=re.IGNORECASE)
            if not (code_m and qty_m and dim_m):
                continue
            m_code = code_m.group("code")
            m_qty = qty_m.group("qty")
            m_dim = dim_m.group("dim")
            dims_nums = re.split(sep, m_dim)
        else:
            m_code = m.group("code")
            m_qty = m.group("qty")
            m_dim = m.group("dim")
            dims_nums = re.split(sep, m_dim)

        # quantidade
        try:
            qty = float(m_qty.replace(",", "."))
        except Exception:
            qty = 0.0

        # dimensões
        larg = comp = esp = 0
        try:
            if len(dims_nums) >= 2:
                larg = int(dims_nums[0])
                comp = int(dims_nums[1])
                if len(dims_nums) >= 3 and dims_nums[2].isdigit():
                    esp = int(dims_nums[2])
        except Exception:
            pass

        # densidade (se houver)
        densidade = ""
        dm = re.search(r"(D\d{2})", line, flags=re.IGNORECASE)
        if dm:
            densidade = dm.group(1).upper()

        produto = {
            "codigo_fornecedor": m_code.upper(),
            "descricao": line,
            "linha_raw": raw,
            "dimensoes": {
                "comprimento": comp,
                "largura": larg,
                "espessura": esp
            },
            "quantidade": qty,
            "unidade": "UNI",
            "mini_codigo": "",  # calculado já de seguida
        }
        produto["mini_codigo"] = generate_mini_codigo(produto)
        products.append(produto)

    return products


def generate_mini_codigo(linha):
    """Gera Mini Código tolerante a dados parciais (usa densidade se existir)."""
    dims = linha.get("dimensoes", {})
    codigo = linha.get("codigo_fornecedor", "")

    comp = dims.get("comprimento", 0)
    larg = dims.get("largura", 0)
    esp = dims.get("espessura", 0)

    dens_m = re.search(r"(D\d{2})", codigo, flags=re.IGNORECASE)
    densidade = (dens_m.group(1).upper() if dens_m else "")

    if larg and comp and esp:
        core = f"{larg}x{comp}x{esp}"
    elif larg and comp:
        core = f"{larg}x{comp}"
    else:
        core = ""

    if densidade and core:
        return f"{densidade}-{core}"
    return core or codigo


def get_realistic_fallback():
    """Fallback realista (não usado se OCR funcionar)."""
    return {
        "numero_requisicao":
        "REQ-2025-0045",
        "document_number":
        "GR-2025-0234",
        "po_number":
        "GR-2025-0234",
        "supplier_name":
        "Blocos Portugal SA",
        "delivery_date":
        "25/09/2025",
        "lines": [
            {
                "supplier_code": "BLC-D25-200x300x150",
                "description": "Bloco betão celular D25 200x300x150",
                "unit": "UNI",
                "qty": 48,
                "mini_codigo": "D25-200x300x150",
                "dimensoes": {
                    "comprimento": 300,
                    "largura": 200,
                    "espessura": 150
                },
            },
            {
                "supplier_code": "BLC-D30-200x600x200",
                "description": "Bloco betão celular D30 200x600x200",
                "unit": "UNI",
                "qty": 24,
                "mini_codigo": "D30-200x600x200",
                "dimensoes": {
                    "comprimento": 600,
                    "largura": 200,
                    "espessura": 200
                },
            },
        ],
        "totals": {
            "total_lines": 2,
            "total_quantity": 72
        },
    }


# ----------------- Mapeamento + Matching + Export -----------------


def map_supplier_codes(supplier, payload):
    mapped = []

    # Suporta novo formato com 'produtos' (Guia de Remessa extraída)
    if "produtos" in payload and payload["produtos"]:
        for produto in payload["produtos"]:
            # Extrair código do fornecedor da referência de ordem (ex: "1ECWH Nº 10874/25EU" -> "1ECWH")
            referencia = produto.get("referencia_ordem", "")
            supplier_code = referencia.split(" ")[0] if referencia else ""
            
            # Artigo/SKU do produto (garantir que não é None)
            article_code = produto.get("artigo") or ""
            
            # IMPORTANTE: Lookup usando article_code, não supplier_code!
            # supplier_code (ex:"1ECWH") é igual para todas as linhas deste fornecedor
            # article_code (ex:"E0748001901") é único por produto
            mapping = CodeMapping.objects.filter(
                supplier=supplier, supplier_code=article_code).first() if article_code else None
            mapped.append({
                "supplier_code": supplier_code or "",
                "article_code": article_code or "UNKNOWN",
                "description": produto.get("descricao") or "",
                "unit": produto.get("unidade") or "UN",
                "qty": produto.get("quantidade") or 0,
                "internal_sku": (mapping.internal_sku if mapping else None),
                "confidence": (mapping.confidence if mapping else 0.0),
                "po_number_extracted": (produto.get("numero_encomenda") or ""),  # Protege contra None
            })
    # Formato antigo com 'lines' (no formato antigo, supplier_code era o SKU do produto)
    elif "lines" in payload:
        for l in payload.get("lines", []):
            supplier_code = l.get("supplier_code")
            mapping = CodeMapping.objects.filter(
                supplier=supplier, supplier_code=supplier_code).first()
            mapped.append({
                **l,
                "article_code": supplier_code,  # No formato antigo, supplier_code era o artigo/SKU
                "internal_sku": (mapping.internal_sku if mapping else None),
                "confidence": (mapping.confidence if mapping else 0.0),
                "po_number_extracted": "",  # Formato antigo não tem este campo
            })

    return mapped


@transaction.atomic
def create_po_from_nota_encomenda(inbound: InboundDocument, payload: dict):
    """
    Cria automaticamente PurchaseOrder + POLines a partir de uma Nota de Encomenda (Fatura).
    Se o documento tem múltiplas encomendas, cria uma PO separada para cada.
    Extrai: número encomenda, fornecedor, produtos, quantidades, dimensões.
    """
    from .models import PurchaseOrder, POLine
    from collections import defaultdict
    
    # Extrair produtos do payload (suporta formatos: produtos ou lines)
    produtos = payload.get("produtos", [])
    if not produtos:
        produtos = payload.get("lines", [])
    
    # Agrupar produtos por numero_encomenda
    produtos_por_po = defaultdict(list)
    
    for produto in produtos:
        # Extrair numero_encomenda do produto (se existir)
        # Protege contra None: (None or "") retorna ""
        po_number = (produto.get("numero_encomenda") or "").strip()
        
        # Se não tem numero_encomenda no produto, usar fallback do documento
        if not po_number:
            po_number = payload.get("document_number") or payload.get("po_number") or f"PO-{inbound.number}"
        
        produtos_por_po[po_number].append(produto)
    
    # Criar uma PO para cada grupo de produtos
    pos_criadas = []
    primeira_po = None
    
    for po_number, produtos_grupo in produtos_por_po.items():
        # Verificar se PO já existe (evitar duplicados)
        existing_po = PurchaseOrder.objects.filter(number=po_number).first()
        if existing_po:
            print(f"⚠️ PO {po_number} já existe, vinculando produtos à PO existente")
            po = existing_po
        else:
            # Criar nova Purchase Order
            po = PurchaseOrder.objects.create(
                number=po_number,
                supplier=inbound.supplier
            )
            print(f"✅ Criada PO {po_number} para fornecedor {inbound.supplier.name}")
        
        # Criar POLines para cada produto deste grupo
        lines_created = 0
        for produto in produtos_grupo:
            # Extrair dados do produto (garantir que não são None)
            article_code = produto.get("artigo") or produto.get("codigo") or produto.get("supplier_code") or ""
            description = produto.get("descricao") or produto.get("description") or ""
            unit = produto.get("unidade") or produto.get("unit") or "UN"
            qty_ordered = Decimal(str(produto.get("quantidade") or produto.get("qty") or 0))
            
            if not article_code or qty_ordered <= 0:
                continue
            
            # Usar get_or_create para evitar duplicados - se SKU já existe, agregar quantidade
            po_line, created = POLine.objects.get_or_create(
                po=po,
                internal_sku=article_code,
                defaults={
                    'description': description,
                    'unit': unit,
                    'qty_ordered': qty_ordered,
                    'tolerance': 0
                }
            )
            
            if not created:
                # Linha já existia - somar quantidades (ambos são Decimal agora)
                po_line.qty_ordered += qty_ordered
                po_line.save()
                print(f"📊 Agregado {qty_ordered} {unit} ao produto {article_code} na PO {po_number} (total: {po_line.qty_ordered})")
            
            lines_created += 1
        
        print(f"✅ Criadas {lines_created} linhas na PO {po_number}")
        pos_criadas.append(po)
        
        if not primeira_po:
            primeira_po = po
    
    # Vincular documento à primeira PO criada (ou última se todas já existiam)
    if pos_criadas:
        inbound.po = primeira_po
        inbound.save()
        print(f"📎 Documento vinculado à PO {primeira_po.number}")
    
    # Retornar primeira PO (mantém compatibilidade com código existente)
    return primeira_po if primeira_po else None


@transaction.atomic
def process_inbound(inbound: InboundDocument):
    # Estratégia: OCR primeiro, LLM só como fallback se OCR falhar
    # 1. Primeiro: OCR rápido para obter texto e produtos
    ocr_payload = real_ocr_extract(inbound.file.path)
    ocr_text = ocr_payload.get('texto_completo', '')
    ocr_produtos = len(ocr_payload.get('produtos', [])) + len(ocr_payload.get('lines', []))
    
    # 2. Verificar se OCR extraiu produtos com sucesso
    if ocr_produtos > 0:
        # OCR funcionou - usar dados OCR
        print(f"✅ Usando dados do OCR cascade ({ocr_produtos} produtos extraídos)")
        payload = ocr_payload
    else:
        # OCR falhou (0 produtos) - tentar LLM como fallback
        print("⚠️ OCR retornou 0 produtos - tentando LLM como fallback...")
        ollama_data = ollama_extract_document(inbound.file.path, ocr_text=ocr_text)
        
        if ollama_data and ollama_data.get('produtos'):
            # Ollama extraiu dados - usar como fallback
            print(f"✅ LLM fallback: {len(ollama_data.get('produtos', []))} produtos extraídos")
            
            # Converter formato Ollama para formato esperado
            payload = {
                "fornecedor": ollama_data.get('fornecedor', ''),
                "nif": ollama_data.get('nif', ''),
                "numero_documento": ollama_data.get('numero_documento', ''),
                "data_documento": ollama_data.get('data_documento', ''),
                "po_number": ollama_data.get('numero_encomenda', ''),
                "iban": ollama_data.get('iban', ''),
                "produtos": [
                    {
                        "artigo": p.get('codigo', ''),
                        "descricao": p.get('descricao', ''),
                        "quantidade": float(p.get('quantidade') or 0),
                        "preco_unitario": float(p.get('preco_unitario') or 0),
                        "total": float(p.get('total') or 0),
                        "numero_encomenda": p.get('numero_encomenda', '')
                    }
                    for p in ollama_data.get('produtos', [])
                ],
                "total": ollama_data.get('total_geral', 0),
                "texto_completo": ocr_text or "Dados extraídos via Ollama LLM",
                "tipo_documento": ollama_data.get('tipo_documento', 'UNKNOWN_LLM'),
                "extraction_method": "ollama_llm"
            }
        else:
            # LLM também falhou - usar dados OCR mesmo vazios
            print("❌ LLM também falhou - usando payload OCR")
            payload = ocr_payload

    if payload.get("error"):
        ExceptionTask.objects.create(
            inbound=inbound,
            line_ref="OCR",
            issue=f"OCR extraction failed: {payload['error']}")

    inbound.parsed_payload = payload
    inbound.save()

    # Se for Nota de Encomenda (FT), criar PurchaseOrder
    if inbound.doc_type == 'FT':
        print(f"📋 Processando Nota de Encomenda: {inbound.number}")
        po = create_po_from_nota_encomenda(inbound, payload)
        print(f"✅ PO {po.number if po else 'N/A'} criada")
        # CONTINUA para criar linhas de receção também!

    # Validar se ficheiro é ilegível (apenas para Guias de Remessa)
    texto_extraido = payload.get("texto_completo", "")
    produtos_extraidos = payload.get("produtos", [])
    linhas_extraidas = payload.get("lines", [])
    
    # Considerar ficheiro ilegível se:
    # - Texto muito curto (<100 chars)
    # - Documento é guia/fatura mas 0 produtos extraídos
    doc_type = payload.get("tipo_documento", "")
    is_document_with_products = any(x in doc_type for x in ["FATURA", "GUIA"])
    
    if len(texto_extraido) < 100:
        ExceptionTask.objects.create(
            inbound=inbound,
            line_ref="OCR",
            issue="Ficheiro ilegível - texto extraído muito curto (menos de 100 caracteres)"
        )
    elif is_document_with_products and not produtos_extraidos and not linhas_extraidas:
        ExceptionTask.objects.create(
            inbound=inbound,
            line_ref="OCR",
            issue="Ficheiro ilegível - nenhum produto foi extraído do documento"
        )
    
    # Se texto pdfplumber era curto E nenhum produto foi extraído → ficheiro ilegível
    if payload.get("baixa_qualidade_texto") and not produtos_extraidos and not linhas_extraidas:
        ExceptionTask.objects.create(
            inbound=inbound,
            line_ref="OCR",
            issue="Ficheiro ilegível - qualidade de imagem muito baixa (texto quase vazio mesmo após OCR)"
        )
    
    # Validar qualidade dos produtos extraídos
    if produtos_extraidos:
        # Verificar se produtos têm dados válidos
        produtos_invalidos = 0
        for produto in produtos_extraidos:
            codigo = produto.get('artigo') or ''
            descricao = produto.get('descricao') or ''
            quantidade = produto.get('quantidade', 0)
            
            # Produto é VÁLIDO se:
            # - Tem código válido (≥5 chars) E quantidade > 0, OU
            # - Não tem código MAS tem descrição válida (≥10 chars) E quantidade > 0
            tem_codigo_valido = codigo and len(str(codigo)) >= 5
            tem_descricao_valida = descricao and len(str(descricao)) >= 10
            tem_quantidade_valida = quantidade > 0
            
            produto_valido = (tem_codigo_valido or tem_descricao_valida) and tem_quantidade_valida
            
            if not produto_valido:
                produtos_invalidos += 1
        
        # Se >50% dos produtos são inválidos, ficheiro está desformatado
        taxa_invalidos = produtos_invalidos / len(produtos_extraidos)
        if taxa_invalidos > 0.5:
            ExceptionTask.objects.create(
                inbound=inbound,
                line_ref="OCR",
                issue=f"Ficheiro desformatado - {produtos_invalidos}/{len(produtos_extraidos)} produtos com dados inválidos (sem código E sem descrição, ou quantidades zero)"
            )

    # criar linhas de receção
    inbound.lines.all().delete()
    mapped_lines = map_supplier_codes(inbound.supplier, payload)
    for ml in mapped_lines:
        ReceiptLine.objects.create(
            inbound=inbound,
            supplier_code=ml["supplier_code"],
            article_code=ml.get("article_code", ""),
            maybe_internal_sku=ml.get("internal_sku") or "",
            description=ml.get("description", ""),
            unit=ml.get("unit", "UN"),
            qty_received=ml.get("qty", 0),
            po_number_extracted=ml.get("po_number_extracted", ""),  # Armazenar numero_encomenda
        )

    # ===== VINCULAÇÃO DE PO (ANTES DE QUALQUER EXCEÇÃO/MATCHING) =====
    # NOTA: Se doc_type == 'FT', já criou e vinculou PO acima, não desvincular!
    if not inbound.po:
        po_number = payload.get("po_number") or payload.get("document_number")
        if po_number:
            po = PurchaseOrder.objects.filter(number=po_number).first()
            if po:
                inbound.po = po
                inbound.save()
                print(f"🔗 PO vinculada: {po.number}")

    # ===== MATCHING =====
    ok = 0
    issues = 0
    exceptions = []
    
    if inbound.doc_type == 'FT':
        print("📋 Nota de Encomenda: SKIP matching (PO criada, aguarda guias de remessa)")
        doc_items = payload.get("produtos", payload.get("lines", []))
        ok = len(doc_items)
    elif inbound.doc_type == 'GR':
        from .models import POLine
        
        for r in inbound.lines.all():
            # Buscar PO correta usando po_number_extracted da linha (se múltiplas POs)
            target_po = inbound.po  # PO padrão vinculada ao documento
            
            if r.po_number_extracted:
                # Tentar encontrar PO específica para este produto
                specific_po = PurchaseOrder.objects.filter(number=r.po_number_extracted).first()
                if specific_po:
                    target_po = specific_po
                    print(f"🔍 Produto {r.article_code} → PO específica {specific_po.number}")
                else:
                    print(f"⚠️ PO {r.po_number_extracted} não encontrada para produto {r.article_code}, usando PO padrão")
            
            # Se ainda não temos PO, criar exceção
            if not target_po:
                issues += 1
                exceptions.append({
                    "line": r.article_code,
                    "issue": f"PO não encontrada (extraído: {r.po_number_extracted or 'N/A'})",
                })
                continue
            
            mapping = CodeMapping.objects.filter(
                supplier=inbound.supplier,
                supplier_code=r.article_code
            ).first()
            
            if not mapping:
                qty_ordered = float(r.qty_received) if r.qty_received else 0.0
                mapping = CodeMapping.objects.create(
                    supplier=inbound.supplier,
                    supplier_code=r.article_code,
                    internal_sku=r.article_code,
                    qty_ordered=qty_ordered,
                    confidence=0.5
                )
                print(f"🆕 CodeMapping criado automaticamente: {r.article_code} → {r.article_code} (qty: {qty_ordered})")
            
            internal_sku = mapping.internal_sku
            po_line = POLine.objects.filter(po=target_po, internal_sku=internal_sku).first()
            
            if not po_line:
                issues += 1
                exceptions.append({
                    "line": r.article_code,
                    "issue": f"Produto {internal_sku} não encontrado na PO {target_po.number}",
                })
                continue
            
            qty_ordered = float(po_line.qty_ordered)
            qty_already_received = float(po_line.qty_received)
            qty_new = float(r.qty_received)
            qty_total_received = qty_already_received + qty_new
            
            if qty_total_received > qty_ordered:
                issues += 1
                exceptions.append({
                    "line": r.article_code,
                    "issue": f"Quantidade excedida: recebida {qty_total_received} vs pedida {qty_ordered} (PO {target_po.number})",
                })
                continue
            
            po_line.qty_received = qty_total_received
            po_line.save()
            print(f"✅ {internal_sku} (PO {target_po.number}): recebida {qty_new}, total {qty_total_received}/{qty_ordered}")
            
            ok += 1

    res, _ = MatchResult.objects.get_or_create(inbound=inbound)

    # Suporta ambos os formatos (produtos ou lines)
    doc_items = payload.get("produtos", payload.get("lines", []))
    total_lines_in_doc = len(doc_items)
    lines_read_successfully = ok
    first_error_line = None
    if exceptions:
        for idx, item in enumerate(doc_items, 1):
            # Tenta ambos os campos (artigo para produtos, supplier_code para lines)
            item_code = item.get("artigo") or item.get("supplier_code") or ""
            if item_code and any(item_code in ex.get("line", "") for ex in exceptions):
                first_error_line = idx
                break

    # Verificar se há erros de OCR/parsing (falhas críticas de processamento)
    ocr_errors_exist = inbound.exceptions.filter(line_ref="OCR").exists()
    
    # Deletar apenas exceções de matching antigas (preservar exceções de OCR)
    inbound.exceptions.exclude(line_ref="OCR").delete()
    
    # Criar novas exceções de matching
    for ex in exceptions:
        ExceptionTask.objects.create(inbound=inbound,
                                     line_ref=ex["line"],
                                     issue=ex["issue"])
    
    # Definir status baseado no tipo de problema:
    # - error: falha no OCR/parsing (ficheiro ilegível, OCR falhou, etc.)
    # - exceptions: problemas no matching (divergências, SKU não encontrado)
    # - matched: tudo OK
    if ocr_errors_exist:
        res.status = "error"
    else:
        res.status = "matched" if issues == 0 else "exceptions"
    
    res.summary = {
        "lines_ok": ok,
        "lines_issues": issues,
        "total_lines_in_document": total_lines_in_doc,
        "lines_read_successfully": lines_read_successfully,
        "first_error_line": first_error_line,
        "last_successful_line": (lines_read_successfully or None),
    }
    res.certified_id = hashlib.sha256(
        (str(inbound.id) + str(payload)).encode()).hexdigest()[:16]
    res.save()

    return res


def extract_dimensions_from_text(text: str) -> str:
    """Extrai dimensões de uma descrição usando regex.
    
    Procura padrões como: 150x200, 1980x0880x0030, 135 x 190, etc.
    Suporta 2-4 dígitos por dimensão.
    Retorna a dimensão encontrada ou string vazia.
    """
    if not text:
        return ""
    
    import re
    
    patterns = [
        r'(\d{2,4})\s*[xX×]\s*(\d{2,4})\s*[xX×]\s*(\d{1,4})',
        r'(\d{2,4})\s*[xX×]\s*(\d{2,4})',
    ]
    
    for pattern in patterns:
        match = re.search(pattern, text)
        if match:
            return 'x'.join(match.groups())
    
    return ""


def export_document_to_excel(inbound_id: int) -> HttpResponse:
    """Exporta para Excel no formato pedido (Mini Código, Dimensões, Quantidade)."""
    from .models import MiniCodigo
    
    inbound = InboundDocument.objects.get(id=inbound_id)

    wb = Workbook()
    ws = wb.active
    ws.title = "Requisição Processada"

    headers = [
        "Mini Código", "Dimensões (LxCxE)", "Quantidade"
    ]
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="FF6B35",
                              end_color="FF6B35",
                              fill_type="solid")

    for col, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=col, value=h)
        c.font = header_font
        c.fill = header_fill
        c.alignment = Alignment(horizontal="center")

    for row, linha in enumerate(inbound.lines.all(), 2):
        dimensoes = ""
        mini_codigo_from_payload = ""
        descricao = ""
        article_code_from_doc = linha.article_code

        # Formato Guia de Remessa (novo): procura em 'produtos' usando article_code
        if inbound.parsed_payload.get("produtos"):
            for produto in inbound.parsed_payload.get("produtos", []):
                # Match usando article_code (código do produto único)
                if produto.get("artigo") == linha.article_code:
                    dims = produto.get("dimensoes", "")
                    # dimensoes pode ser string (Tesseract) ou dicionário (formato antigo)
                    if isinstance(dims, str):
                        dimensoes = dims
                    elif isinstance(dims, dict) and any(dims.values()):
                        larg = dims.get("largura", 0)
                        comp = dims.get("comprimento", 0)
                        esp = dims.get("espessura", 0)
                        if larg and comp and esp:
                            dimensoes = f"{larg}x{comp}x{esp}"
                        elif larg and comp:
                            dimensoes = f"{larg}x{comp}"
                    mini_codigo_from_payload = produto.get("mini_codigo", "")
                    descricao = produto.get("descricao", "")
                    break
        
        # Formato antigo: procura em 'lines' usando supplier_code
        else:
            for payload_line in inbound.parsed_payload.get("lines", []):
                if payload_line.get("supplier_code") == linha.supplier_code:
                    dims = payload_line.get("dimensoes", "")
                    # dimensoes pode ser string (Tesseract) ou dicionário (formato antigo)
                    if isinstance(dims, str):
                        dimensoes = dims
                    elif isinstance(dims, dict) and any(dims.values()):
                        larg = dims.get("largura", 0)
                        comp = dims.get("comprimento", 0)
                        esp = dims.get("espessura", 0)
                        if larg and comp and esp:
                            dimensoes = f"{larg}x{comp}x{esp}"
                        elif larg and comp:
                            dimensoes = f"{larg}x{comp}"
                    mini_codigo_from_payload = payload_line.get("mini_codigo", "")
                    descricao = payload_line.get("description", "")
                    break

        # Fallback: se não houver dimensões, tenta extrair da descrição
        if not dimensoes:
            dimensoes = extract_dimensions_from_text(descricao or linha.description)
        
        # 🎯 PRIORIDADE 1: MAPEAR MINI CÓDIGO DA BASE DE DADOS
        # Tenta mapear usando article_code → identificador na BD
        mini_codigo_from_db = None
        if article_code_from_doc:
            try:
                mini_obj = MiniCodigo.objects.filter(identificador=article_code_from_doc).first()
                if mini_obj:
                    mini_codigo_from_db = mini_obj.mini_codigo
                    # Se não temos designação do documento, usa da BD
                    if not descricao:
                        descricao = mini_obj.designacao
            except Exception as e:
                print(f"⚠️ Erro ao mapear mini código por article_code: {e}")
        
        # Fallback: tenta mapear usando supplier_code se article_code não funcionou
        if not mini_codigo_from_db and linha.supplier_code:
            try:
                mini_obj = MiniCodigo.objects.filter(identificador=linha.supplier_code).first()
                if mini_obj:
                    mini_codigo_from_db = mini_obj.mini_codigo
                    if not descricao:
                        descricao = mini_obj.designacao
            except Exception as e:
                print(f"⚠️ Erro ao mapear mini código por supplier_code: {e}")
        
        # Hierarquia de fallback: BD → payload → maybe_internal_sku → article_code
        final_mini_codigo = (
            mini_codigo_from_db or 
            mini_codigo_from_payload or 
            linha.maybe_internal_sku or 
            article_code_from_doc
        )

        ws.cell(row=row, column=1, value=final_mini_codigo)
        ws.cell(row=row, column=2, value=dimensoes)
        ws.cell(row=row, column=3, value=float(linha.qty_received))

    # auto width
    for column in ws.columns:
        max_len = 0
        letter = column[0].column_letter
        for cell in column:
            try:
                max_len = max(max_len, len(str(cell.value)))
            except Exception:
                pass
        ws.column_dimensions[letter].width = min(max_len + 2, 50)

    response = HttpResponse(
        content_type=
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response[
        "Content-Disposition"] = f'attachment; filename="requisicao_{inbound.id}.xlsx"'
    wb.save(response)
    return response
